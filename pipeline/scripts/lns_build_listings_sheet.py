"""
build_listings_sheet.py  —  Adds "Properties For Sale" sheet to LNS_Real_Estate_Agents_FULL.xlsx
Reads listings JSON from stdin, cross-references agent phones from lns_agents_full.json,
adds a sheet to the existing Excel workbook.
"""
import json, sys, re, glob
from pathlib import Path
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

# ── path discovery ─────────────────────────────────────────────────────────────
_candidates = glob.glob('/sessions/*/mnt/Downloads/lns_agents_full.json')
if _candidates:
    BASE = Path(_candidates[0]).parent
else:
    _dirs = glob.glob('/sessions/*/mnt/Downloads')
    BASE = Path(_dirs[0]) if _dirs else (Path.home() / 'Downloads')

AGENTS_JSON = BASE / 'lns_agents_full.json'
XL_PATH     = BASE / 'LNS_Real_Estate_Agents_FULL.xlsx'
SHEET_NAME  = "🏠 Properties For Sale"

# ── load agent database for phone cross-reference ──────────────────────────────
print(f"Loading agent database from {AGENTS_JSON}...")
with open(AGENTS_JSON) as f:
    agents_raw = json.load(f)

# Build lookup: lowercase agent name → {mobile, telephone}
agent_phone_map = {}
for a in agents_raw:
    name_key = (a.get('name') or '').strip().lower()
    if name_key:
        agent_phone_map[name_key] = {
            'mobile':    a.get('mobile', ''),
            'telephone': a.get('telephone', ''),
        }
print(f"  {len(agent_phone_map)} agents in database for cross-reference")

# ── helpers ────────────────────────────────────────────────────────────────────
def fmt_phone(p):
    if not p: return ''
    d = re.sub(r'\D', '', str(p))
    if len(d) == 10:
        return f"{d[:4]} {d[4:7]} {d[7:]}" if d.startswith('04') else f"{d[:2]} {d[2:6]} {d[6:]}"
    return str(p).strip()

def lookup_agent_phones(agent_names_str):
    """Cross-reference agent names with our database to find phone numbers."""
    if not agent_names_str:
        return '', ''
    names = [n.strip() for n in re.split(r'[,/]', agent_names_str)]
    mobiles, phones = [], []
    for name in names:
        info = agent_phone_map.get(name.lower(), {})
        m = fmt_phone(info.get('mobile', ''))
        t = fmt_phone(info.get('telephone', ''))
        if m: mobiles.append(m)
        if t and t not in phones: phones.append(t)
    return ' / '.join(mobiles[:2]), ' / '.join(phones[:1])

def parse_price_num(price_str):
    """Extract a rough numeric price for sorting (returns 0 if can't parse)."""
    if not price_str: return 0
    nums = re.findall(r'\d[\d,.]*', price_str)
    for n in nums:
        try:
            v = float(n.replace(',', ''))
            if v > 100000:
                return v
        except: pass
    return 0

def fmt_land(size, unit):
    if not size or size == 0: return ''
    return f"{int(size):,} {unit}"

def fmt_auction(d):
    if not d: return ''
    try:
        from datetime import datetime
        dt = datetime.fromisoformat(d.replace('Z','+00:00'))
        return dt.strftime('%-d %b %Y %I:%M%p').replace('AM','am').replace('PM','pm')
    except: return d

def tag_label(tag, listing_type):
    if listing_type == 'project': return '🏗 Development'
    if not tag: return ''
    tl = tag.lower()
    if tl == 'new': return '🆕 New'
    if tl == 'new home': return '🏗 New Home'
    if tl == 'updated': return '📝 Updated'
    if 'price reduced' in tl or 'price drop' in tl: return '⬇ Price Reduced'
    if 'under offer' in tl or 'under contract' in tl: return '🤝 Under Offer'
    if tl.startswith('auction'): return '🔨 ' + tag  # e.g. "Auction Wed 1 Apr"
    return tag

# ── styles ─────────────────────────────────────────────────────────────────────
NAVY  = '1A3A5C'; TEAL = '0D7377'; WHITE = 'FFFFFF'
LIGHT = 'EBF2FA'; LGREY = 'F9FAFB'; GREEN = 'E8F5E9'
AMBER = 'FFF8E1'; RED_LIGHT = 'FFEBEE'; BLUE_LIGHT = 'E3F2FD'

NAVY_FILL  = PatternFill("solid", fgColor=NAVY)
TEAL_FILL  = PatternFill("solid", fgColor=TEAL)
ALT_FILL   = PatternFill("solid", fgColor=LIGHT)
GREEN_FILL = PatternFill("solid", fgColor=GREEN)
AMBER_FILL = PatternFill("solid", fgColor=AMBER)
RED_FILL   = PatternFill("solid", fgColor=RED_LIGHT)
BLUE_FILL  = PatternFill("solid", fgColor=BLUE_LIGHT)

HDR_FONT  = Font(name="Arial", bold=True, color=WHITE, size=10)
ROW_FONT  = Font(name="Arial", size=9)
BOLD_FONT = Font(name="Arial", bold=True, size=9)
LINK_FONT = Font(name="Arial", size=9, color="1565C0", underline="single")
SUB_FONT  = Font(name="Arial", bold=True, color=WHITE, size=9)
THIN = Side(style='thin', color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR = Alignment(horizontal='center', vertical='center', wrap_text=False)
LFT = Alignment(horizontal='left',   vertical='center', wrap_text=False)

# ── column definitions ─────────────────────────────────────────────────────────
COLS = [
    ('Status',         14, CTR),
    ('Address',        38, LFT),
    ('Suburb',         16, LFT),
    ('Type',           14, CTR),
    ('Beds',            6, CTR),
    ('Baths',           6, CTR),
    ('Cars',            6, CTR),
    ('Land (m²)',      10, CTR),
    ('Price Guide',    30, LFT),
    ('Auction Date',   16, CTR),
    ('Agency',         28, LFT),
    ('Agents',         30, LFT),
    ('Agent Mobile',   16, CTR),
    ('Agent Phone',    14, CTR),
    ('Price Change',   16, CTR),
    ('Domain Link',    14, CTR),
]
N = len(COLS)

# ── main ───────────────────────────────────────────────────────────────────────
def build(listings):
    today = date.today().strftime("%-d %b %Y")
    print(f"Processing {len(listings)} listings...")

    # Enrich each listing with phone lookups and sort order
    rows = []
    suburb_order = [
        'Mosman','Neutral Bay','Cremorne','Cremorne Point','Kirribilli',
        'Mcmahons Point','Lavender Bay','Milsons Point','Cammeray',
        'Crows Nest','North Sydney','Waverton','Wollstonecraft','Kurraba Point',
        'Naremburn','St Leonards','Artarmon','Greenwich','Longueville'
    ]
    suburb_rank = {s: i for i, s in enumerate(suburb_order)}

    for l in listings:
        mobile, phone = lookup_agent_phones(l.get('agentNames', ''))
        # address field already contains full address e.g. "3 Smith St Mosman NSW"
        # Strip trailing suburb/state if address ends with them to avoid double suburb
        raw_addr = l.get('address', '')
        sub = l.get('suburb', '')
        # Clean address: remove trailing "SuburbName NSW" etc
        addr_clean = re.sub(r'\s+' + re.escape(sub) + r'\s+NSW\s*$', '', raw_addr, flags=re.IGNORECASE).strip()
        addr_clean = re.sub(r'\s+NSW\s*$', '', addr_clean, flags=re.IGNORECASE).strip()
        price_num = parse_price_num(l.get('price',''))
        tag = l.get('tagText', '')
        rows.append({
            'Status':       tag_label(tag, ''),
            'Address':      addr_clean if addr_clean else raw_addr,
            'Suburb':       sub,
            'Type':         l.get('propertyType',''),
            'Beds':         l.get('beds',0) or '',
            'Baths':        l.get('baths',0) or '',
            'Cars':         l.get('parking',0) or '',
            'Land (m²)':    l.get('landSize',0) or '',
            'Price Guide':  l.get('price',''),
            'Auction Date': fmt_auction(l.get('auctionDate','')),
            'Agency':       l.get('agencyName',''),
            'Agents':       l.get('agentNames',''),
            'Agent Mobile': mobile,
            'Agent Phone':  phone,
            'Price Change': ('⬇ Price Reduced' if 'price reduced' in (tag or '').lower()
                           else ('📝 Updated'    if (tag or '').lower() == 'updated'
                           else ('🆕 New'        if (tag or '').lower() in ('new','new home')
                           else ('🤝 Under Offer' if 'under offer' in (tag or '').lower()
                           else ('🔨 Auction'    if (tag or '').lower().startswith('auction')
                           else ''))))),
            'Domain Link':  l.get('url',''),
            '_suburb_rank': suburb_rank.get(sub, 99),  # type: ignore
            '_price_num':   price_num,
            '_tag':         l.get('tag',''),
        })

    # Sort: suburb order, then price descending
    rows.sort(key=lambda r: (r['_suburb_rank'], -r['_price_num']))

    # ── load workbook and add/replace sheet ────────────────────────────────────
    print(f"Loading workbook {XL_PATH}...")
    wb = load_workbook(XL_PATH)

    # Remove old sheet if exists
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    # Insert as second sheet (after All Agents)
    ws = wb.create_sheet(SHEET_NAME, 1)
    ws.freeze_panes = "B3"

    # ── title row ──────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(N)}1")
    tc = ws['A1']
    tc.value = (f"LNS Properties For Sale — domain.com.au  |  {len(rows)} listings  |  "
                f"Updated {today}  |  ⬇ Price Reduced = Reduced  |  📝 Updated = Changed")
    tc.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    tc.fill = TEAL_FILL; tc.alignment = CTR
    ws.row_dimensions[1].height = 20

    # ── header row ─────────────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 26
    for ci, (key, w, al) in enumerate(COLS, 1):
        c = ws.cell(2, ci, key)
        c.font = HDR_FONT; c.fill = NAVY_FILL; c.alignment = CTR; c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── data rows with suburb separators ──────────────────────────────────────
    current_suburb = None; row_num = 3
    col_keys = [c[0] for c in COLS]
    col_aligns = [c[2] for c in COLS]

    # Tag → fill colour
    PURPLE_FILL = PatternFill("solid", fgColor="F3E5F5")
    INDIGO_FILL = PatternFill("solid", fgColor="E8EAF6")

    def get_row_fill(status, alt):
        if '⬇' in status: return RED_FILL
        if '📝' in status: return AMBER_FILL
        if '🆕' in status: return GREEN_FILL
        if '🏗' in status: return BLUE_FILL
        if '🤝' in status: return PURPLE_FILL
        if '🔨' in status: return INDIGO_FILL
        return ALT_FILL if alt else None

    for r in rows:
        # Suburb separator
        if r['Suburb'] != current_suburb:
            current_suburb = r['Suburb']
            sub_count = sum(1 for x in rows if x['Suburb'] == current_suburb)
            ws.merge_cells(f"A{row_num}:{get_column_letter(N)}{row_num}")
            sc = ws.cell(row_num, 1)
            sc.value = f"  {current_suburb.upper()}  —  {sub_count} listings"
            sc.font = SUB_FONT; sc.fill = TEAL_FILL; sc.alignment = LFT
            ws.row_dimensions[row_num].height = 15; row_num += 1

        # Row fill based on status
        row_fill = get_row_fill(r['Status'], row_num % 2 == 0)

        for ci, key in enumerate(col_keys, 1):
            val = r.get(key, '')
            cell = ws.cell(row_num, ci, val)
            cell.border = BORDER; cell.font = ROW_FONT; cell.alignment = col_aligns[ci-1]
            if row_fill: cell.fill = row_fill

            if key == 'Domain Link' and val:
                short = val.split('domain.com.au')[-1][:30] + '…' if len(val.split('domain.com.au')[-1]) > 30 else val.split('domain.com.au')[-1]
                cell.value = short; cell.hyperlink = val; cell.font = LINK_FONT
            elif key == 'Price Guide':
                cell.font = Font(name="Arial", bold=True, size=9) if val else ROW_FONT
            elif key == 'Price Change' and '⬇' in str(val):
                cell.font = Font(name="Arial", bold=True, size=9, color="C62828")
            elif key in ('Agent Mobile', 'Agent Phone') and val:
                cell.font = Font(name="Arial", size=9, color="1565C0")
            elif key == 'Status':
                s = r['Status']
                if '⬇' in s:   cell.font = Font(name="Arial", bold=True, size=9, color="C62828")
                elif '📝' in s: cell.font = Font(name="Arial", bold=True, size=9, color="E65100")
                elif '🆕' in s: cell.font = Font(name="Arial", bold=True, size=9, color="2E7D32")
                elif '🏗' in s: cell.font = Font(name="Arial", bold=True, size=9, color="1565C0")
                elif '🤝' in s: cell.font = Font(name="Arial", bold=True, size=9, color="6A1B9A")
                elif '🔨' in s: cell.font = Font(name="Arial", bold=True, size=9, color="283593")
                else:            cell.font = ROW_FONT

        ws.row_dimensions[row_num].height = 15
        row_num += 1

    # ── summary stats at bottom ────────────────────────────────────────────────
    row_num += 1
    ws.merge_cells(f"A{row_num}:{get_column_letter(N)}{ row_num}")
    stat_cell = ws.cell(row_num, 1)
    reduced     = sum(1 for r in rows if '⬇' in r.get('Price Change',''))
    updated     = sum(1 for r in rows if '📝' in r.get('Price Change',''))
    new_c       = sum(1 for r in rows if '🆕' in r.get('Status',''))
    under_offer = sum(1 for r in rows if '🤝' in r.get('Status',''))
    auction     = sum(1 for r in rows if '🔨' in r.get('Status',''))
    stat_cell.value = (f"Total: {len(rows)} listings  |  ⬇ Price Reduced: {reduced}  |  "
                       f"📝 Updated: {updated}  |  🆕 New: {new_c}  |  "
                       f"🤝 Under Offer: {under_offer}  |  🔨 Auction: {auction}  |  Source: domain.com.au")
    stat_cell.font = Font(name="Arial", italic=True, size=9, color="555555")
    stat_cell.alignment = CTR

    wb.save(XL_PATH)
    print(f"  Sheet '{SHEET_NAME}' saved to {XL_PATH}")
    print(f"  {len(rows)} listings  |  {reduced} price reduced  |  {updated} updated  |  {new_c} new")

if __name__ == '__main__':
    data = sys.stdin.read().strip()
    listings = json.loads(data)
    build(listings)

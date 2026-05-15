"""
lns_build_offmarket_sheet.py
Inserts "🏘 Off Market" sheet between "🏠 Properties For Sale" and "🏡 Sold Properties".
Sources: onthehouse.com.au listings not found in Domain for-sale list.
"""
import json, re, glob
from pathlib import Path
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── path discovery ──────────────────────────────────────────────────────────
_c = glob.glob('/sessions/*/mnt/Downloads/lns_agents_full.json')
BASE = Path(_c[0]).parent if _c else Path.home() / 'Downloads'

XL_PATH       = BASE / 'LNS_Real_Estate_Agents_FULL.xlsx'
AGENTS_JSON   = BASE / 'lns_agents_full.json'
OFFMKT_JSON   = BASE / 'lns_offmarket_raw.json'
FORSALE_JSON  = BASE / 'lns_listings_raw.json'
SHEET_NAME    = "🏘 Off Market"

SUBURB_ORDER = [
    'Mosman','Neutral Bay','Cremorne','Cremorne Point','Kirribilli',
    'Mcmahons Point','Lavender Bay','Milsons Point','Cammeray',
    'Crows Nest','North Sydney','Waverton','Wollstonecraft','Kurraba Point',
    'Naremburn','St Leonards','Artarmon','Greenwich','Longueville'
]

# ── styles ───────────────────────────────────────────────────────────────────
NAVY   = '1A3A5C'; TEAL  = '0D7377'; WHITE = 'FFFFFF'
LIGHT  = 'EBF2FA'; PURPLE= '4A148C'; AMBER = 'E65100'
NAVY_FILL   = PatternFill("solid", fgColor=NAVY)
TEAL_FILL   = PatternFill("solid", fgColor=TEAL)
PURPLE_FILL = PatternFill("solid", fgColor=PURPLE)
AMBER_FILL  = PatternFill("solid", fgColor=AMBER)
ALT_FILL    = PatternFill("solid", fgColor=LIGHT)
HDR_FONT    = Font(name="Arial", bold=True, color=WHITE, size=10)
ROW_FONT    = Font(name="Arial", size=9)
BOLD_FONT   = Font(name="Arial", bold=True, size=9)
LINK_FONT   = Font(name="Arial", size=9, color="1565C0", underline="single")
THIN   = Side(style='thin', color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR = Alignment(horizontal='center', vertical='center')
LFT = Alignment(horizontal='left',   vertical='center')

COLS = [
    ('Address',       36, LFT),
    ('Suburb',        16, LFT),
    ('Type',          14, CTR),
    ('Beds',           7, CTR),
    ('Baths',          7, CTR),
    ('Cars',           7, CTR),
    ('Land m²',       10, CTR),
    ('Price Guide',   22, LFT),
    ('Agency',        26, LFT),
    ('Agent Mobile',  16, CTR),
    ('Agent Phone',   16, CTR),
    ('Source',        12, CTR),
    ('OTH Link',      14, CTR),
]
N_COLS = len(COLS)

def fmt_phone(p):
    if not p: return ''
    d = re.sub(r'\D', '', str(p))
    if len(d) == 10:
        return f"{d[:4]} {d[4:7]} {d[7:]}" if d.startswith('04') else f"{d[:2]} {d[2:6]} {d[6:]}"
    return str(p).strip()

# ── load agent phone map ─────────────────────────────────────────────────────
with open(AGENTS_JSON) as f:
    agents_raw = json.load(f)
agent_phone_map = {}
for a in agents_raw:
    key = (a.get('agencyName') or '').strip().lower()
    if key and key not in agent_phone_map:
        agent_phone_map[key] = {
            'mobile': fmt_phone(a.get('mobile','')),
            'telephone': fmt_phone(a.get('telephone',''))
        }

def agency_phones(agency_name):
    key = (agency_name or '').strip().lower()
    info = agent_phone_map.get(key, {})
    return info.get('mobile',''), info.get('telephone','')

# ── load off-market listings (raw + manual entries) ──────────────────────────
with open(OFFMKT_JSON) as f:
    offmkt = json.load(f)

# Merge in manual entries from "My Off Markets" sheet if present
MANUAL_JSON = BASE / 'lns_offmarket_manual.json'
if MANUAL_JSON.exists():
    with open(MANUAL_JSON) as f:
        manual = json.load(f)
    if manual:
        # Deduplicate: skip manual entries already in offmkt by normalised address
        ABBREV2 = {'street':'st','avenue':'ave','road':'rd','crescent':'cres',
                   'place':'pl','drive':'dr','close':'cl','lane':'ln','court':'ct'}
        def _norm(s):
            s = re.sub(r'[\s,./]+', ' ', str(s).lower().strip())
            for f2, a2 in ABBREV2.items():
                s = re.sub(r'\b'+f2+r'\b', a2, s)
            return s
        existing_norms = set(_norm(l.get('address','') + ' ' + l.get('suburb','')) for l in offmkt)
        added = 0
        for m in manual:
            mn = _norm(m.get('address','') + ' ' + m.get('suburb',''))
            if mn not in existing_norms:
                offmkt.append({
                    'address': m.get('address',''),
                    'suburb':  m.get('suburb',''),
                    'type':    m.get('type',''),
                    'beds':    m.get('beds',''),
                    'baths':   m.get('baths',''),
                    'cars':    m.get('cars',''),
                    'land':    m.get('land',''),
                    'price':   m.get('price',''),
                    'agency':  m.get('agency',''),
                    'link':    '',
                    'source':  m.get('source','Manual'),
                })
                existing_norms.add(mn)
                added += 1
        if added:
            print(f"  + {added} manual entries merged from 'My Off Markets' sheet")

print(f"  Off-market listings loaded: {len(offmkt)}")

# Sort by suburb order then address
suburb_rank = {s: i for i, s in enumerate(SUBURB_ORDER)}
offmkt_sorted = sorted(offmkt,
    key=lambda l: (suburb_rank.get(l.get('suburb',''), 99), l.get('address','')))

def build_offmarket_sheet(wb, today_str):
    # Remove existing sheet if present
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    # Insert after "🏠 Properties For Sale"
    forsale_idx = wb.sheetnames.index('🏠 Properties For Sale') if '🏠 Properties For Sale' in wb.sheetnames else 1
    ws = wb.create_sheet(SHEET_NAME, forsale_idx + 1)

    # Column widths
    for ci, (key, w, _) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Title row
    ws.merge_cells(f'A1:{get_column_letter(N_COLS)}1')
    tc = ws['A1']
    tc.value = (f"🏘  OFF MARKET PROPERTIES  —  Lower North Shore  |  "
                f"{len(offmkt_sorted)} listings not on Domain/REA  |  "
                f"Source: onthehouse.com.au  |  Updated {today_str}")
    tc.font  = Font(name="Arial", bold=True, size=11, color=WHITE)
    tc.fill  = PURPLE_FILL
    tc.alignment = CTR
    ws.row_dimensions[1].height = 22

    # Header row
    ws.row_dimensions[2].height = 26
    for ci, (key, w, al) in enumerate(COLS, 1):
        c = ws.cell(2, ci, key)
        c.font = HDR_FONT; c.fill = NAVY_FILL
        c.alignment = CTR; c.border = BORDER

    ws.freeze_panes = "A3"

    row_num = 3
    current_suburb = None

    for l in offmkt_sorted:
        suburb = l.get('suburb', '')

        # Suburb separator
        if suburb != current_suburb:
            current_suburb = suburb
            ws.merge_cells(f'A{row_num}:{get_column_letter(N_COLS)}{row_num}')
            sc = ws.cell(row_num, 1)
            sc.value    = f"  {suburb.upper()}"
            sc.font     = Font(name="Arial", bold=True, color=WHITE, size=9)
            sc.fill     = TEAL_FILL
            sc.alignment = LFT
            ws.row_dimensions[row_num].height = 16
            row_num += 1

        alt = ALT_FILL if row_num % 2 == 0 else None
        mob, pho = agency_phones(l.get('agency',''))

        # Format price text
        price_raw = (l.get('price') or '').strip()
        if price_raw.lower() in ('for sale', ''):
            price_display = 'Contact Agent'
        else:
            price_display = price_raw.title() if price_raw.islower() else price_raw

        values = {
            'Address':      l.get('address',''),
            'Suburb':       suburb,
            'Type':         l.get('type',''),
            'Beds':         l.get('beds','') or '',
            'Baths':        l.get('baths','') or '',
            'Cars':         l.get('cars','') or '',
            'Land m²':      l.get('land','') or '',
            'Price Guide':  price_display,
            'Agency':       l.get('agency',''),
            'Agent Mobile': mob,
            'Agent Phone':  pho,
            'Source':       'OnTheHouse',
            'OTH Link':     l.get('link',''),
        }

        for ci, (key, _, al) in enumerate(COLS, 1):
            val = values.get(key, '')
            c = ws.cell(row_num, ci, val)
            c.border = BORDER; c.font = ROW_FONT; c.alignment = al
            if alt: c.fill = alt
            if key == 'OTH Link' and val:
                short = '…' + val.split('onthehouse.com.au')[-1][-30:]
                c.value = short; c.hyperlink = val; c.font = LINK_FONT
            elif key == 'Price Guide' and val and val != 'Contact Agent':
                c.font = Font(name="Arial", bold=True, size=9, color="1A3A5C")
            elif key in ('Agent Mobile', 'Agent Phone') and val:
                c.font = Font(name="Arial", size=9, color="1565C0")
            elif key == 'Address':
                c.font = BOLD_FONT
        ws.row_dimensions[row_num].height = 15
        row_num += 1

    # Footer
    row_num += 1
    ws.merge_cells(f'A{row_num}:{get_column_letter(N_COLS)}{row_num}')
    ws.cell(row_num, 1).value = (
        f"Total: {len(offmkt_sorted)} off-market properties  |  "
        f"Source: onthehouse.com.au (properties not listed on Domain or REA)  |  "
        f"Updated {today_str}"
    )
    ws.cell(row_num, 1).font      = Font(name="Arial", italic=True, size=9, color="555555")
    ws.cell(row_num, 1).alignment = CTR

    print(f"  '{SHEET_NAME}' sheet built: {len(offmkt_sorted)} off-market listings")

# ── MAIN ─────────────────────────────────────────────────────────────────────
today_str = date.today().strftime('%-d %b %Y')
print(f"Loading workbook {XL_PATH}...")
wb = load_workbook(XL_PATH)
print(f"  Existing sheets: {wb.sheetnames}")

build_offmarket_sheet(wb, today_str)

wb.save(XL_PATH)
print(f"\nSaved → {XL_PATH}")
print(f"  Sheets: {wb.sheetnames}")

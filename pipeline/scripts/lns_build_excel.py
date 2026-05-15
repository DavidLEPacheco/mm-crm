"""
lns_build_excel.py  —  LNS Real Estate Agent spreadsheet builder
Reads lns_agents_full.json → writes LNS_Real_Estate_Agents_FULL.xlsx
Automatically discovers the correct mnt/Downloads path for the current session.
Run: python3 lns_build_excel.py
"""
import json, re, sys, glob
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# Auto-discover the mnt Downloads path
_candidates = glob.glob('/sessions/*/mnt/Downloads/lns_agents_full.json')
if _candidates:
    BASE = Path(_candidates[0]).parent
else:
    # Fallback: look for any mnt/Downloads
    _dirs = glob.glob('/sessions/*/mnt/Downloads')
    BASE = Path(_dirs[0]) if _dirs else Path.home() / 'Downloads'

IN_JSON = BASE / 'lns_agents_full.json'
OUT_XL  = BASE / 'LNS_Real_Estate_Agents_FULL.xlsx'

SUBURB_ORDER = [
    'Mosman','Neutral Bay','Cremorne','Cremorne Point','Kirribilli',
    'Mcmahons Point','Lavender Bay','Milsons Point','Cammeray',
    'Crows Nest','North Sydney','Waverton','Wollstonecraft','Kurraba Point',
    'Naremburn','St Leonards','Artarmon','Greenwich','Longueville'
]

def fmt_phone(p):
    if not p: return ''
    d = re.sub(r'\D','',str(p))
    if len(d)==10:
        return f"{d[:4]} {d[4:7]} {d[7:]}" if d.startswith('04') else f"{d[:2]} {d[2:6]} {d[6:]}"
    return str(p).strip()

def fmt_currency(v):
    return f"${int(v):,}" if v else ''

TIER_SCORE = {'platinum':4,'gold':3,'silver':2,'bronze':1}
TIER_LABEL = {'platinum':'⭐ Platinum','gold':'🥇 Gold',
              'silver':'🥈 Silver','bronze':'🥉 Bronze','':'Standard'}

def priority_score(a):
    tier  = TIER_SCORE.get(a.get('profileTier',''),0)*15
    sales = min((a.get('totalSold') or 0)+(a.get('totalJointSold') or 0),100)*0.5
    rat   = (a.get('starRating') or 0)*4
    rev   = min(a.get('numReviews') or 0, 200)*0.05
    return round(tier+sales+rat+rev, 1)

NAVY='1A3A5C'; TEAL='0D7377'; WHITE='FFFFFF'; LIGHT='EBF2FA'; LGREY='F9FAFB'
NAVY_FILL = PatternFill("solid",fgColor=NAVY)
TEAL_FILL = PatternFill("solid",fgColor=TEAL)
ALT_FILL  = PatternFill("solid",fgColor=LIGHT)
ALSO_FILL = PatternFill("solid",fgColor="E8F5E9")
HDR_FONT  = Font(name="Arial",bold=True,color=WHITE,size=10)
ROW_FONT  = Font(name="Arial",size=9)
BOLD_FONT = Font(name="Arial",bold=True,size=9)
LINK_FONT = Font(name="Arial",size=9,color="1565C0",underline="single")
SUB_FONT  = Font(name="Arial",bold=True,color=WHITE,size=9)
THIN      = Side(style='thin',color="CCCCCC")
BORDER    = Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
CTR = Alignment(horizontal='center',vertical='center')
LFT = Alignment(horizontal='left',  vertical='center')

TIER_FILLS = {
    '⭐ Platinum': PatternFill("solid",fgColor="FFF3CD"),
    '🥇 Gold':     PatternFill("solid",fgColor="FFF8E1"),
    '🥈 Silver':   PatternFill("solid",fgColor="F5F5F5"),
    '🥉 Bronze':   PatternFill("solid",fgColor="FBE9E7"),
}
DATE_FILL      = PatternFill("solid", fgColor="FFFDE7")   # pale yellow for Contacted
DATE_FILL_DONE = PatternFill("solid", fgColor="E8F5E9")   # pale green once filled
NOTES_FILL     = PatternFill("solid", fgColor="F8F9FA")   # very light grey for Notes

COLS = [
    ('Call Priority',  10, CTR),
    ('Name',           26, LFT),
    ('Agency',         28, LFT),
    ('Contacted',      14, CTR),   # click cell then Ctrl+; to stamp today's date
    ('Notes',          45, LFT),   # free-text notes from calls
    ('Mobile',         16, CTR),
    ('Agency Website', 28, LFT),
    ('Primary Suburb', 16, LFT),
    ('Also Active In', 32, LFT),
    ('For Sale Now',   12, CTR),
    ('Sold (Solo)',     12, CTR),
    ('Total Sales',    12, CTR),
    ('Domain Profile', 14, CTR),
]
col_keys   = [c[0] for c in COLS]
col_aligns = [c[2] for c in COLS]
N = len(COLS)

def write_sheet(ws, data, title_text, title_fill, suburb_separators=False):
    ws.freeze_panes = "B3"
    ws.merge_cells(f"A1:{get_column_letter(N)}1")
    tc = ws['A1']
    tc.value=title_text; tc.font=Font(name="Arial",bold=True,size=11,color=WHITE)
    tc.fill=title_fill; tc.alignment=CTR
    ws.row_dimensions[1].height=20
    ws.row_dimensions[2].height=26
    for ci,(key,w,al) in enumerate(COLS,1):
        label = key
        if key == 'Contacted': label = 'Contacted\n(Ctrl+; = today)'
        if key == 'Notes':     label = 'Notes  (call feedback)'
        c=ws.cell(2,ci,label); c.font=HDR_FONT
        c.fill = PatternFill("solid",fgColor="E65100") if key=='Contacted' else \
                 PatternFill("solid",fgColor="1565C0") if key=='Notes' else NAVY_FILL
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        c.border=BORDER
        ws.column_dimensions[get_column_letter(ci)].width=w

    current_suburb=None; row_num=3
    for agent in data:
        if suburb_separators and agent['Primary Suburb']!=current_suburb:
            current_suburb=agent['Primary Suburb']
            ws.merge_cells(f"A{row_num}:{get_column_letter(N)}{row_num}")
            sc=ws.cell(row_num,1); sc.value=f"  {current_suburb.upper()}"
            sc.font=SUB_FONT; sc.fill=TEAL_FILL; sc.alignment=LFT
            ws.row_dimensions[row_num].height=16; row_num+=1

        alt=(row_num%2==0); has_also=bool(agent.get('Also Active In',''))
        for ci,key in enumerate(col_keys,1):
            val=agent.get(key,'')
            cell=ws.cell(row_num,ci,val)
            cell.border=BORDER; cell.font=ROW_FONT; cell.alignment=col_aligns[ci-1]
            # Contacted and Notes always get their own fills regardless of row alt
            if key == 'Contacted':
                cell.fill = DATE_FILL_DONE if val else DATE_FILL
                cell.font = Font(name="Arial",size=9,bold=True,color="2E7D32") if val else \
                            Font(name="Arial",size=8,italic=True,color="AAAAAA")
                if not val: cell.value = ''   # leave blank for user to fill
                cell.alignment = CTR
            elif key == 'Notes':
                cell.fill = NOTES_FILL
                cell.alignment = Alignment(horizontal='left',vertical='center',wrap_text=True)
            elif has_also:
                cell.fill=ALSO_FILL
            elif alt:
                cell.fill=ALT_FILL
            if key not in ('Contacted','Notes') and not has_also:
                tf=TIER_FILLS.get(agent.get('Tier',''))
                if tf and key in ('Name','Call Priority'): cell.fill=tf
            if key=='Domain Profile' and val:
                cell.value=val.split('/real-estate-agent/')[-1].rstrip('/')
                cell.hyperlink=val; cell.font=LINK_FONT
            elif key=='Call Priority':
                v=val or 0
                cell.font=Font(name="Arial",bold=True,size=9,
                    color="2E7D32" if v>=50 else("F0A500" if v>=30 else "555555"))
            elif key=='Total Sales':  cell.font=BOLD_FONT
            elif key=='Also Active In' and val:
                cell.font=Font(name="Arial",size=9,italic=True,color="1B5E20")
        ws.row_dimensions[row_num].height=16; row_num+=1

    ts_col = get_column_letter(col_keys.index('Total Sales')+1)
    ws.conditional_formatting.add(f"{ts_col}3:{ts_col}{row_num}",
        ColorScaleRule(start_type='num',start_value=0, start_color='FFFFFF',
                       mid_type='num', mid_value=30,   mid_color='FFF176',
                       end_type='num', end_value=100,  end_color='1B5E20'))
    ws.conditional_formatting.add(f"A3:A{row_num}",
        ColorScaleRule(start_type='num',start_value=0, start_color='FFCDD2',
                       mid_type='num', mid_value=40,   mid_color='FFF9C4',
                       end_type='num', end_value=80,   end_color='C8E6C9'))

def build(raw):
    from datetime import date
    today = date.today().strftime("%-d %b %Y")

    agents=[]
    seen=set()
    for s in SUBURB_ORDER:
        grp=sorted(
            [a for a in raw if a.get('suburb')==s and a['id'] not in seen],
            key=lambda x:-priority_score(x)
        )
        for a in grp:
            seen.add(a['id'])
            total_sales=(a.get('totalSold') or 0)+(a.get('totalJointSold') or 0)
            also_in_raw = [x for x in (a.get('allSuburbs') or []) if x!=s]
            agents.append({
                'Call Priority':  priority_score(a),
                'Name':           a['name'].strip(),
                'Agency':         (a.get('agencyName') or '').strip(),
                'Contacted':      '',   # user fills in date with Ctrl+;
                'Notes':          '',   # user fills in call notes
                'Mobile':         fmt_phone(a.get('mobile','')),
                'Agency Website': (a.get('agencyWebsite') or '').strip(),
                'Primary Suburb': s,
                'Also Active In': ', '.join(also_in_raw),
                'For Sale Now':   a.get('totalForSale') or 0,
                'Sold (Solo)':    a.get('totalSold') or 0,
                'Total Sales':    total_sales,
                'Domain Profile': a.get('profileUrl',''),
            })

    multi=[a for a in agents if a['Also Active In']]
    print(f"  {len(agents)} agents, {len(multi)} active across multiple suburbs")

    wb=Workbook()
    ws=wb.active; ws.title="All Agents"
    write_sheet(ws, agents,
        f"Lower North Shore — {len(agents)} Real Estate Agents  |  domain.com.au  |  Updated {today}  |  🟢 Green = active across multiple suburbs",
        NAVY_FILL, suburb_separators=True)

    ws2=wb.create_sheet("🔥 Top Performers")
    top=sorted(agents,key=lambda x:-x['Call Priority'])
    write_sheet(ws2, top,
        f"TOP PERFORMERS — by Call Priority  |  Updated {today}  |  🟢 Green = active across multiple suburbs",
        TEAL_FILL, suburb_separators=False)

    ws3=wb.create_sheet("Summary")
    ws3.column_dimensions['A'].width=28; ws3.column_dimensions['B'].width=14
    ws3.column_dimensions['D'].width=24; ws3.column_dimensions['E'].width=12

    def sh(r,c,v,bold=False,sz=10,color="000000",fill=None,align=LFT,bdr=True):
        cell=ws3.cell(r,c,v)
        cell.font=Font(name="Arial",bold=bold,size=sz,color=color)
        if fill: cell.fill=fill
        cell.alignment=align
        if bdr: cell.border=BORDER
        return cell

    sh(1,1,f"LNS Real Estate Agents — Updated {today}",bold=True,sz=13,color=NAVY,bdr=False)
    ws3.merge_cells("A1:E1")
    sh(2,1,f"Source: domain.com.au  |  {len(agents)} unique agents  |  {len(multi)} active across multiple suburbs",sz=9,color="555555",bdr=False)
    ws3.merge_cells("A2:E2")

    sh(4,1,"Suburb",bold=True,fill=NAVY_FILL,color=WHITE,align=CTR)
    sh(4,2,"Agents",bold=True,fill=NAVY_FILL,color=WHITE,align=CTR)
    sh(4,4,"Agency",bold=True,fill=NAVY_FILL,color=WHITE,align=CTR)
    sh(4,5,"Agents",bold=True,fill=NAVY_FILL,color=WHITE,align=CTR)
    suburb_counts={}
    for a in agents: suburb_counts[a['Primary Suburb']]=suburb_counts.get(a['Primary Suburb'],0)+1
    for i,s in enumerate(SUBURB_ORDER,5):
        f=ALT_FILL if i%2==0 else None
        sh(i,1,s,fill=f); sh(i,2,suburb_counts.get(s,0),fill=f,align=CTR)
    agency_counts={}
    for a in agents: agency_counts[a['Agency']]=agency_counts.get(a['Agency'],0)+1
    for i,(ag,n) in enumerate(sorted(agency_counts.items(),key=lambda x:-x[1])[:15],5):
        f=ALT_FILL if i%2==0 else None
        sh(i,4,ag,fill=f,align=LFT); sh(i,5,n,fill=f,align=CTR)

    sh(28,1,f"Agents active across multiple suburbs ({len(multi)})",bold=True,sz=10,color=NAVY,bdr=False)
    ws3.merge_cells("A28:E28")
    sh(29,1,"Agent",bold=True,fill=TEAL_FILL,color=WHITE)
    sh(29,2,"Primary Suburb",bold=True,fill=TEAL_FILL,color=WHITE)
    sh(29,4,"Also Active In",bold=True,fill=TEAL_FILL,color=WHITE)
    ws3.merge_cells("D29:E29")
    for i,a in enumerate(sorted(multi,key=lambda x:x['Name']),30):
        f=ALT_FILL if i%2==0 else None
        sh(i,1,a['Name'],fill=f); sh(i,2,a['Primary Suburb'],fill=f,align=CTR)
        sh(i,4,a['Also Active In'],fill=f); ws3.merge_cells(f"D{i}:E{i}")

    wb.save(OUT_XL)
    print(f"  Saved → {OUT_XL}")
    return len(agents), len(multi)

if __name__=='__main__':
    print(f"Loading {IN_JSON}...")
    with open(IN_JSON) as f:
        raw=json.load(f)
    n,m=build(raw)
    print(f"Done. {n} agents, {m} multi-suburb.")

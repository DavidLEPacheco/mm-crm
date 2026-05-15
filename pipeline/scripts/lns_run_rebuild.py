"""
lns_run_rebuild.py  —  Master orchestrator for LNS workbook rebuild
Runs all build scripts in the correct order, then fixes sheet ordering.

Usage:
    python3 lns_run_rebuild.py           # full rebuild (all sheets)
    python3 lns_run_rebuild.py --buyers  # buyers sheet only (fast, after updating clients)
"""
import sys, glob, subprocess, json
from pathlib import Path

# ── path discovery ──────────────────────────────────────────────────────────
_c = glob.glob('/sessions/*/mnt/Downloads/lns_agents_full.json')
BASE    = Path(_c[0]).parent if _c else Path.home() / 'Downloads'
SCRIPTS = BASE / 'lns_agents_scripts'
XL_PATH = BASE / 'LNS_Real_Estate_Agents_FULL.xlsx'

LISTINGS_JSON = BASE / 'lns_listings_raw.json'
CLIENTS_JSON  = BASE / 'lns_clients.json'

def run(cmd, stdin_file=None, label=''):
    print(f"\n{'='*60}")
    print(f"  {label or ' '.join(cmd)}")
    print(f"{'='*60}")
    if stdin_file and Path(stdin_file).exists():
        with open(stdin_file) as f:
            result = subprocess.run(cmd, stdin=f, capture_output=False)
    else:
        result = subprocess.run(cmd, capture_output=False)
    if result.returncode != 0:
        print(f"  ⚠  Exit code {result.returncode}")
    return result.returncode == 0

def fix_sheet_order():
    """Ensure sheets are in correct display order after rebuild."""
    from openpyxl import load_workbook
    DESIRED = [
        'All Agents',
        '👥 Buyers',
        '🏠 Properties For Sale',
        '🏘 Off Market',
        '📝 My Off Markets',
        '🏡 Sold Properties',
        '🔥 Top Performers',
        'Summary',
    ]
    wb = load_workbook(XL_PATH)
    existing = wb.sheetnames
    target = [s for s in DESIRED if s in existing] + \
             [s for s in existing if s not in DESIRED]
    for final_pos, name in enumerate(target):
        cur = wb.sheetnames.index(name)
        if cur != final_pos:
            wb.move_sheet(name, offset=final_pos - cur)
    wb.save(XL_PATH)
    print(f"\n  ✅ Sheet order fixed: {wb.sheetnames}")

def buyers_only():
    """Rebuild just the Buyers matching sheet (fast path after editing client profiles)."""
    print("\n🔄  Running buyer matching update only...")
    ok = run(['python3', str(SCRIPTS / 'lns_build_buyers_sheet.py')],
             label='Build Buyers Sheet')
    fix_sheet_order()
    if ok:
        print("\n✅  Buyer matches updated successfully!")
    else:
        print("\n⚠  Buyer matching finished with warnings — check output above.")

def full_rebuild():
    """Full workbook rebuild — all sheets in order."""
    print("\n🔄  Starting full LNS workbook rebuild...")

    steps = [
        # 1. Base agents workbook (creates the .xlsx)
        (['python3', str(SCRIPTS / 'lns_build_excel.py')],
         None, '1/6  Build All Agents sheet'),

        # 2. For Sale listings (reads from stdin)
        (['python3', str(SCRIPTS / 'lns_build_listings_sheet.py')],
         str(LISTINGS_JSON) if LISTINGS_JSON.exists() else None,
         '2/6  Build Properties For Sale sheet'),

        # 3. My Off Markets (harvest manual entries + rebuild input sheet)
        (['python3', str(SCRIPTS / 'lns_build_myoffmarket_sheet.py')],
         None, '3/6  Build My Off Markets sheet'),

        # 4. Off Market (auto-scraped + merges manual entries)
        (['python3', str(SCRIPTS / 'lns_build_offmarket_sheet.py')],
         None, '4/6  Build Off Market sheet'),

        # 5. Buyers matching
        (['python3', str(SCRIPTS / 'lns_build_buyers_sheet.py')],
         None, '5/6  Build Buyers sheet'),

        # 6. Sold + Summary
        (['python3', str(SCRIPTS / 'lns_build_sold_and_summary.py')],
         None, '6/6  Build Sold + Summary sheet'),
    ]

    errors = 0
    for cmd, stdin_file, label in steps:
        if not run(cmd, stdin_file=stdin_file, label=label):
            errors += 1

    fix_sheet_order()

    print(f"\n{'='*60}")
    if errors == 0:
        print("  ✅  Full rebuild complete — no errors!")
    else:
        print(f"  ⚠   Rebuild complete with {errors} step(s) that had warnings.")
    print(f"  📄  Workbook: {XL_PATH}")
    print(f"{'='*60}\n")

if __name__ == '__main__':
    if '--buyers' in sys.argv:
        buyers_only()
    else:
        full_rebuild()

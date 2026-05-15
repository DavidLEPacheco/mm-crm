"""
lns_build_myoffmarket_sheet.py
Manages the "📝 My Off Markets" sheet — a manual-entry input area where
you type your own off-market properties. Entries persist across daily rebuilds.

Flow:
  1. Read any existing entries from the sheet (if it already exists) → save to lns_offmarket_manual.json
  2. Rebuild the sheet with those entries preserved, plus blank rows ready for new ones
  3. The manual entries are merged into lns_offmarket_raw.json by lns_build_offmarket_sheet.py
"""

import json, glob, re
from pathlib import Path
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── path discovery ──────────────────────────────────────────────────────────────
_c = glob.glob('/sessions/*/mnt/Downloads/lns_agents_full.json')
BASE    = Path(_c[0]).parent if _c else Path.home() / 'Downloads'
XL_PATH = BASE / 'LNS_Real_Estate_Agents_FULL.xlsx'
MANUAL_JSON = BASE / 'lns_offmarket_manual.json'
SHEET_NAME  = '📝 My Off Markets'

# ── styles ──────────────────────────────────────────────────────────────────────
PURPLE     = '6A1B9A'
PURPLE_MID = '9C27B0'
PURPLE_LIGHT = 'F3E5F5'
YELLOW_LIGHT = 'FFFDE7'
WHITE  = 'FFFFFF'
LGREY  = 'F5F5F5'
BORDER_CLR = 'CE93D8'

PRP_FILL   = PatternFill("solid", fgColor=PURPLE)
PRP2_FILL  = PatternFill("solid", fgColor=PURPLE_MID)
PLGT_FILL  = PatternFill("solid", fgColor=PURPLE_LIGHT)
YLW_FILL   = PatternFill("solid", fgColor=YELLOW_LIGHT)
GRY_FILL   = PatternFill("solid", fgColor=LGREY)

HDR_FONT  = Font(name='Arial', bold=True, color=WHITE, size=10)
ROW_FONT  = Font(name='Arial', size=9)
HINT_FONT = Font(name='Arial', size=8, italic=True, color='9E9E9E')
TITLE_FONT= Font(name='Arial', bold=True, size=12, color=WHITE)

THIN  = Side(style='thin',   color=BORDER_CLR)
VTHIN = Side(style='thin',   color='E0E0E0')
BDR   = Border(left=THIN,  right=THIN,  top=THIN,  bottom=THIN)
VBDR  = Border(left=VTHIN, right=VTHIN, top=VTHIN, bottom=VTHIN)

CTR = Alignment(horizontal='center', vertical='center', wrap_text=False)
LFT = Alignment(horizontal='left',   vertical='center', wrap_text=False)
WRAP= Alignment(horizontal='left',   vertical='center', wrap_text=True)

# ── column definitions ──────────────────────────────────────────────────────────
COLS = [
    ('Address',        32, LFT),
    ('Suburb',         16, LFT),
    ('Type',           14, CTR),
    ('Beds',            6, CTR),
    ('Baths',           6, CTR),
    ('Cars',            6, CTR),
    ('Land m²',        10, CTR),
    ('Price Guide',    24, LFT),
    ('Agency',         26, LFT),
    ('Agent Mobile',   16, CTR),
    ('Notes / Source', 35, LFT),
]
N = len(COLS)

BLANK_ROWS = 30   # number of empty input rows to provide

LNS_SUBURBS = [
    'Mosman','Neutral Bay','Cremorne','Cremorne Point','Kirribilli',
    'McMahons Point','Lavender Bay','Milsons Point','Cammeray',
    'Crows Nest','North Sydney','Waverton','Wollstonecraft',
    'Naremburn','St Leonards','Artarmon','Greenwich','Longueville'
]

def read_existing_entries(wb):
    """Read manually entered rows from the existing sheet, if present."""
    if SHEET_NAME not in wb.sheetnames:
        return []
    ws = wb[SHEET_NAME]
    entries = []
    col_keys = [c[0] for c in COLS]
    for row in ws.iter_rows(min_row=5, values_only=True):
        # Row has data if Address column is non-empty
        vals = list(row[:N])
        if vals[0] and str(vals[0]).strip() and str(vals[0]).strip() != '← Start typing here':
            entry = {}
            for i, key in enumerate(col_keys):
                v = vals[i] if i < len(vals) else None
                entry[key] = str(v).strip() if v is not None else ''
            entries.append(entry)
    print(f"  Read {len(entries)} existing manual entries from '{SHEET_NAME}'")
    return entries

def save_manual_json(entries):
    """Convert sheet entries to off-market JSON format and save."""
    records = []
    for e in entries:
        if not e.get('Address','').strip():
            continue
        records.append({
            'address': e.get('Address','').strip(),
            'suburb':  e.get('Suburb','').strip(),
            'type':    e.get('Type','').strip(),
            'beds':    e.get('Beds','').strip(),
            'baths':   e.get('Baths','').strip(),
            'cars':    e.get('Cars','').strip(),
            'land':    e.get('Land m²','').strip(),
            'price':   e.get('Price Guide','').strip(),
            'agency':  e.get('Agency','').strip(),
            'agent_mobile': e.get('Agent Mobile','').strip(),
            'link':    '',
            'source':  'Manual — ' + (e.get('Notes / Source','').strip() or 'Own entry'),
        })
    with open(MANUAL_JSON, 'w') as f:
        json.dump(records, f, indent=2)
    print(f"  Saved {len(records)} manual entries → {MANUAL_JSON.name}")
    return records

def build_sheet(wb, entries):
    """Build/replace the My Off Markets sheet."""
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    today = date.today().strftime('%-d %b %Y')

    # Insert after Off Market sheet, or as 3rd sheet
    insert_pos = 3
    for i, name in enumerate(wb.sheetnames):
        if name == '🏘 Off Market':
            insert_pos = i + 1
            break

    ws = wb.create_sheet(SHEET_NAME, insert_pos)
    ws.freeze_panes = 'A5'
    ws.sheet_view.showGridLines = False

    # ── Row 1: Title ────────────────────────────────────────────────────────────
    ws.merge_cells(f'A1:{get_column_letter(N)}1')
    tc = ws['A1']
    tc.value = f'📝  My Off Market Properties  —  Enter your own listings here  |  Updated {today}'
    tc.font  = TITLE_FONT
    tc.fill  = PRP_FILL
    tc.alignment = CTR
    ws.row_dimensions[1].height = 24

    # ── Row 2: Instructions ─────────────────────────────────────────────────────
    ws.merge_cells(f'A2:{get_column_letter(N)}2')
    ic = ws['A2']
    ic.value = ('ℹ  Type properties you know about directly into this table.  '
                'They are automatically merged into the Off Market tab when the workbook rebuilds.  '
                'Your entries are saved and never overwritten by the daily update.')
    ic.font  = Font(name='Arial', size=9, italic=True, color='4A148C')
    ic.fill  = PLGT_FILL
    ic.alignment = WRAP
    ws.row_dimensions[2].height = 32

    # ── Row 3: Blank spacer ─────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ── Row 4: Column headers ───────────────────────────────────────────────────
    ws.row_dimensions[4].height = 26
    for ci, (key, w, al) in enumerate(COLS, 1):
        c = ws.cell(4, ci, key)
        c.font  = HDR_FONT
        c.fill  = PRP2_FILL
        c.alignment = CTR
        c.border = BDR
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Data rows: existing entries first, then blank rows ─────────────────────
    col_keys = [c[0] for c in COLS]
    total_rows = max(len(entries) + BLANK_ROWS, BLANK_ROWS)

    for ri in range(total_rows):
        row_num = ri + 5
        is_entry = ri < len(entries)
        fill = YLW_FILL if is_entry else (PLGT_FILL if ri % 2 == 0 else None)
        ws.row_dimensions[row_num].height = 18

        for ci, (key, _, al) in enumerate(COLS, 1):
            c = ws.cell(row_num, ci)
            c.font   = ROW_FONT
            c.border = VBDR
            c.alignment = al

            if is_entry:
                c.value = entries[ri].get(key, '')
                c.fill  = YLW_FILL
            else:
                # First blank row: subtle placeholder hint in Address cell
                if ci == 1 and ri == len(entries):
                    c.value = '← Start typing here'
                    c.font  = HINT_FONT
                if fill:
                    c.fill = fill

    # ── Status bar at bottom ────────────────────────────────────────────────────
    bot_row = total_rows + 5 + 2
    ws.merge_cells(f'A{bot_row}:{get_column_letter(N)}{bot_row}')
    sc = ws[f'A{bot_row}']
    sc.value = (f'✅  {len(entries)} manual entries saved  |  '
                f'{BLANK_ROWS} blank rows available  |  '
                f'These entries appear in the 🏘 Off Market tab after each rebuild')
    sc.font  = Font(name='Arial', size=9, italic=True, color=PURPLE)
    sc.fill  = PLGT_FILL
    sc.alignment = LFT

    print(f"  '📝 My Off Markets' sheet built: {len(entries)} existing entries + {BLANK_ROWS} blank rows")

def run():
    print(f'Loading workbook {XL_PATH}...')
    wb = load_workbook(XL_PATH)
    print(f'  Existing sheets: {wb.sheetnames}')

    # Step 1: harvest any existing entries
    entries = read_existing_entries(wb)

    # Step 2: save to JSON for off-market merge
    save_manual_json(entries)

    # Step 3: rebuild the sheet
    build_sheet(wb, entries)

    wb.save(XL_PATH)
    print(f'\nSaved → {XL_PATH}')
    print(f'  Sheets: {load_workbook(XL_PATH, read_only=True).sheetnames}')

if __name__ == '__main__':
    run()

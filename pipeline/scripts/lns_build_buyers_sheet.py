"""
lns_build_buyers_sheet.py
Manages the "👥 Buyers" sheet:

  TOP SECTION  — Client profile table you fill in once (preserved across rebuilds)
  BOTTOM SECTION — Auto-matched properties from both For Sale and Off Market tabs

Client criteria matched:
  • Budget min / max  (price guide parsed to numeric)
  • Min bedrooms
  • Preferred suburbs  (comma-separated list)
  • Property type      (House / Apartment / Townhouse / Any)
"""

import json, glob, re
from pathlib import Path
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── path discovery ──────────────────────────────────────────────────────────────
_c = glob.glob('/sessions/*/mnt/Downloads/lns_agents_full.json')
BASE        = Path(_c[0]).parent if _c else Path.home() / 'Downloads'
XL_PATH     = BASE / 'LNS_Real_Estate_Agents_FULL.xlsx'
CLIENTS_JSON= BASE / 'lns_clients.json'
SHEET_NAME  = '👥 Buyers'

# ── styles ──────────────────────────────────────────────────────────────────────
TEAL     = '00695C'
TEAL_MID = '00897B'
TEAL_LGT = 'E0F2F1'
NAVY     = '1A3A5C'
AMBER    = 'FF8F00'
AMBER_L  = 'FFF8E1'
GREEN    = '2E7D32'
GREEN_L  = 'E8F5E9'
MATCH_BLUE = 'E3F2FD'
MATCH_GRN  = 'F1F8E9'
WHITE    = 'FFFFFF'
LGREY    = 'F9FAFB'
RED_L    = 'FFEBEE'

TEAL_FILL  = PatternFill("solid", fgColor=TEAL)
TEAL2_FILL = PatternFill("solid", fgColor=TEAL_MID)
TEAL_L_FILL= PatternFill("solid", fgColor=TEAL_LGT)
NAVY_FILL  = PatternFill("solid", fgColor=NAVY)
AMB_FILL   = PatternFill("solid", fgColor=AMBER_L)
GRN_FILL   = PatternFill("solid", fgColor=GREEN_L)
BLU_FILL   = PatternFill("solid", fgColor=MATCH_BLUE)
GRN2_FILL  = PatternFill("solid", fgColor=MATCH_GRN)
GRY_FILL   = PatternFill("solid", fgColor=LGREY)

HDR_FONT   = Font(name='Arial', bold=True, color=WHITE,    size=10)
ROW_FONT   = Font(name='Arial', size=9)
BOLD_FONT  = Font(name='Arial', bold=True, size=9)
TITLE_FONT = Font(name='Arial', bold=True, size=12, color=WHITE)
HINT_FONT  = Font(name='Arial', size=8, italic=True, color='9E9E9E')
CLIENT_HDR = Font(name='Arial', bold=True, size=10, color=WHITE)
MATCH_FONT = Font(name='Arial', size=9)

THIN  = Side(style='thin',   color='B2DFDB')
VTHIN = Side(style='thin',   color='E0E0E0')
THICK = Side(style='medium', color=TEAL)
BDR   = Border(left=THIN,  right=THIN,  top=THIN,  bottom=THIN)
VBDR  = Border(left=VTHIN, right=VTHIN, top=VTHIN, bottom=VTHIN)
TBDR  = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

CTR  = Alignment(horizontal='center', vertical='center', wrap_text=False)
LFT  = Alignment(horizontal='left',   vertical='center', wrap_text=False)
WRAP = Alignment(horizontal='left',   vertical='center', wrap_text=True)

# ── client profile columns ──────────────────────────────────────────────────────
PROFILE_COLS = [
    ('Client Name',        18, LFT),
    ('Budget Min $',       14, CTR),
    ('Budget Max $',       14, CTR),
    ('Min Beds',            8, CTR),
    ('Property Type',      16, CTR),   # House / Apartment / Townhouse / Any
    ('Preferred Suburbs',  38, LFT),   # comma-separated
    ('Notes',              32, LFT),
]
PC = len(PROFILE_COLS)
MAX_CLIENTS = 15   # rows in profile table

# ── match result columns ────────────────────────────────────────────────────────
MATCH_COLS = [
    ('Client',        18, LFT),
    ('Source',        14, CTR),
    ('Address',       30, LFT),
    ('Suburb',        16, LFT),
    ('Type',          12, CTR),
    ('Beds',           6, CTR),
    ('Baths',          6, CTR),
    ('Price Guide',   26, LFT),
    ('Agency',        26, LFT),
    ('Agents',        28, LFT),
    ('Agent Mobile',  16, CTR),
]
MC = len(MATCH_COLS)

# ── price parsing ───────────────────────────────────────────────────────────────
def parse_price(price_str):
    """Extract best-guess numeric price from a price guide string."""
    if not price_str:
        return None
    s = str(price_str).lower()
    if 'contact' in s or 'eoi' in s or 'expression' in s:
        return None
    nums = re.findall(r'\$?([\d,]+(?:\.\d+)?)\s*[mk]?', s)
    values = []
    for n in nums:
        try:
            v = float(n.replace(',',''))
            # Handle shorthand: 1.5m → 1500000, 900k → 900000
            if re.search(r'\b' + re.escape(n.replace(',','')) + r'\s*m\b', s):
                v *= 1_000_000
            elif re.search(r'\b' + re.escape(n.replace(',','')) + r'\s*k\b', s):
                v *= 1_000
            if v > 50_000:
                values.append(v)
        except:
            pass
    if not values:
        return None
    # Return midpoint of range if multiple values
    return sum(values) / len(values)

def price_in_budget(price_str, budget_min, budget_max):
    """Check if property price falls within budget. None budget = no constraint."""
    if budget_min is None and budget_max is None:
        return True
    p = parse_price(price_str)
    if p is None:
        return True   # "Contact Agent" → always show (can't filter)
    if budget_min is not None and p < budget_min:
        return False
    if budget_max is not None and p > budget_max:
        return False
    return True

def suburbs_match(prop_suburb, pref_suburbs):
    """True if property suburb is in client's preferred list (or list is empty)."""
    if not pref_suburbs:
        return True
    prop_s = prop_suburb.lower().strip()
    for ps in pref_suburbs:
        if ps.lower().strip() in prop_s or prop_s in ps.lower().strip():
            return True
    return False

def type_match(prop_type, client_type):
    """True if property type matches client preference (or client has no preference)."""
    if not client_type or client_type.lower() in ('any', 'all', ''):
        return True
    prop_t = (prop_type or '').lower()
    client_t = client_type.lower()
    if not prop_t:
        return True   # unknown type → always show
    return client_t in prop_t or prop_t in client_t

def beds_match(prop_beds, min_beds):
    """True if property has at least min_beds bedrooms."""
    if min_beds is None or min_beds == 0:
        return True
    try:
        return int(prop_beds or 0) >= int(min_beds)
    except:
        return True

# ── read existing client profiles from sheet ────────────────────────────────────
def read_client_profiles(wb):
    """Read client profile rows from existing sheet."""
    if SHEET_NAME not in wb.sheetnames:
        return []
    ws = wb[SHEET_NAME]
    clients = []
    # Profile table starts at row 5, columns A:G
    for row in ws.iter_rows(min_row=5, max_row=5 + MAX_CLIENTS - 1, values_only=True):
        name = row[0]
        if not name or not str(name).strip() or str(name).startswith('e.g.'):
            continue
        def val(idx):
            v = row[idx] if idx < len(row) else None
            return str(v).strip() if v is not None else ''
        def num(idx):
            v = row[idx] if idx < len(row) else None
            try:
                return float(str(v).replace(',','').replace('$','').strip()) if v else None
            except:
                return None

        suburbs = [s.strip() for s in re.split(r'[,;]+', val(5)) if s.strip()]
        clients.append({
            'name':    str(name).strip(),
            'bud_min': num(1),
            'bud_max': num(2),
            'min_beds':num(3),
            'prop_type': val(4),
            'suburbs': suburbs,
            'notes':   val(6),
        })
    print(f"  Read {len(clients)} client profiles from '{SHEET_NAME}'")
    return clients

def save_clients_json(clients):
    with open(CLIENTS_JSON, 'w') as f:
        json.dump(clients, f, indent=2)
    print(f"  Saved {len(clients)} clients → {CLIENTS_JSON.name}")

# ── load listings for matching ───────────────────────────────────────────────────
def load_all_listings(base):
    forsale, offmarket = [], []
    fs_path = base / 'lns_listings_raw.json'
    om_path = base / 'lns_offmarket_raw.json'
    mn_path = base / 'lns_offmarket_manual.json'

    if fs_path.exists():
        with open(fs_path) as f:
            raw = json.load(f)
        for l in raw:
            # Clean address (strip trailing suburb/state)
            addr = l.get('address','')
            sub  = l.get('suburb','')
            addr = re.sub(r'\s+' + re.escape(sub) + r'\s+NSW\s*$', '', addr, flags=re.IGNORECASE).strip()
            addr = re.sub(r'\s+NSW\s*$', '', addr, flags=re.IGNORECASE).strip()
            forsale.append({
                'source':  'For Sale',
                'address': addr,
                'suburb':  sub,
                'type':    l.get('propertyType',''),
                'beds':    l.get('beds') or '',
                'baths':   l.get('baths') or '',
                'price':   l.get('price',''),
                'agency':  l.get('agencyName',''),
                'agents':  l.get('agentNames',''),
            })

    if om_path.exists():
        with open(om_path) as f:
            raw = json.load(f)
        for l in raw:
            offmarket.append({
                'source':  'Off Market',
                'address': l.get('address',''),
                'suburb':  l.get('suburb',''),
                'type':    l.get('type',''),
                'beds':    l.get('beds') or '',
                'baths':   l.get('baths') or '',
                'price':   l.get('price',''),
                'agency':  l.get('agency',''),
                'agents':  '',
            })

    # Also include manual entries
    if mn_path.exists():
        with open(mn_path) as f:
            manual = json.load(f)
        for l in manual:
            offmarket.append({
                'source':  '📝 My Off Market',
                'address': l.get('address',''),
                'suburb':  l.get('suburb',''),
                'type':    l.get('type',''),
                'beds':    l.get('beds') or '',
                'baths':   l.get('baths') or '',
                'price':   l.get('price',''),
                'agency':  l.get('agency',''),
                'agents':  '',
            })

    print(f"  Loaded {len(forsale)} for-sale, {len(offmarket)} off-market listings for matching")
    return forsale, offmarket

def find_matches(clients, forsale, offmarket):
    """Return list of (client_name, listing) tuples for all matches."""
    all_listings = forsale + offmarket
    results = []
    for client in clients:
        for listing in all_listings:
            if (price_in_budget(listing['price'], client['bud_min'], client['bud_max'])
                    and beds_match(listing['beds'], client['min_beds'])
                    and suburbs_match(listing['suburb'], client['suburbs'])
                    and type_match(listing['type'], client['prop_type'])):
                results.append((client['name'], listing))
    return results

# ── agent phone lookup ───────────────────────────────────────────────────────────
def load_agent_phones(base):
    path = base / 'lns_agents_full.json'
    if not path.exists():
        return {}
    with open(path) as f:
        agents = json.load(f)
    lookup = {}
    for a in agents:
        name = (a.get('name') or '').strip().lower()
        if name:
            m = a.get('mobile','')
            d = re.sub(r'\D','',str(m))
            if len(d)==10:
                m = f"{d[:4]} {d[4:7]} {d[7:]}" if d.startswith('04') else f"{d[:2]} {d[2:6]} {d[6:]}"
            lookup[name] = m
    return lookup

# ── build sheet ─────────────────────────────────────────────────────────────────
def build_sheet(wb, clients, matches):
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    # Insert as 2nd sheet (right after All Agents)
    ws = wb.create_sheet(SHEET_NAME, 1)
    ws.freeze_panes = 'A5'
    ws.sheet_view.showGridLines = False

    today = date.today().strftime('%-d %b %Y')
    total_cols = max(PC, MC)

    # ── Row 1: Title ─────────────────────────────────────────────────────────────
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    t = ws['A1']
    t.value = f'👥  Buyer Matching  —  {len(clients)} clients  |  {len(matches)} property matches  |  Updated {today}'
    t.font  = TITLE_FONT
    t.fill  = TEAL_FILL
    t.alignment = CTR
    ws.row_dimensions[1].height = 24

    # ── Row 2: Section header — Client Profiles ──────────────────────────────────
    ws.merge_cells(f'A2:{get_column_letter(PC)}2')
    h = ws['A2']
    h.value = '▌ CLIENT PROFILES  —  Fill in your buyers below. Changes are saved on each rebuild.'
    h.font  = Font(name='Arial', bold=True, size=10, color=WHITE)
    h.fill  = TEAL2_FILL
    h.alignment = LFT
    ws.row_dimensions[2].height = 20

    # ── Row 3: Column headers for profiles ──────────────────────────────────────
    ws.row_dimensions[3].height = 26
    for ci, (key, w, al) in enumerate(PROFILE_COLS, 1):
        c = ws.cell(3, ci, key)
        c.font  = HDR_FONT
        c.fill  = NAVY_FILL
        c.alignment = CTR
        c.border = BDR
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Row 4: Hint row ──────────────────────────────────────────────────────────
    hints = [
        'e.g. Smith Family', '1500000', '2500000', '3',
        'House / Apartment / Any', 'Mosman, Neutral Bay, Cremorne',
        'First home buyer, pre-approved'
    ]
    ws.row_dimensions[4].height = 15
    for ci, hint in enumerate(hints, 1):
        c = ws.cell(4, ci, hint)
        c.font  = HINT_FONT
        c.fill  = TEAL_L_FILL
        c.alignment = LFT if ci in (1,6,7) else CTR
        c.border = VBDR

    # ── Profile data rows ────────────────────────────────────────────────────────
    profile_keys = [k for k,_,_ in PROFILE_COLS]
    client_row_map = {}  # client_name → row number (for colouring match section)

    for ri in range(MAX_CLIENTS):
        rn = ri + 5
        ws.row_dimensions[rn].height = 18
        fill = AMB_FILL if ri % 2 == 0 else GRY_FILL

        if ri < len(clients):
            c_data = clients[ri]
            vals = [
                c_data['name'],
                int(c_data['bud_min']) if c_data['bud_min'] else '',
                int(c_data['bud_max']) if c_data['bud_max'] else '',
                int(c_data['min_beds']) if c_data['min_beds'] else '',
                c_data['prop_type'],
                ', '.join(c_data['suburbs']),
                c_data['notes'],
            ]
            client_row_map[c_data['name']] = ri
        else:
            vals = [''] * PC

        for ci, v in enumerate(vals, 1):
            c = ws.cell(rn, ci, v)
            c.font   = ROW_FONT
            c.fill   = fill
            c.alignment = LFT if ci in (1,6,7) else CTR
            c.border = VBDR
            if ci == 2 or ci == 3:  # Budget columns
                c.number_format = '#,##0'

    # ── Spacer row ───────────────────────────────────────────────────────────────
    sep_row = 5 + MAX_CLIENTS + 1
    ws.row_dimensions[sep_row].height = 8

    # ── Matches section header ───────────────────────────────────────────────────
    match_hdr_row = sep_row + 1
    ws.merge_cells(f'A{match_hdr_row}:{get_column_letter(MC)}{match_hdr_row}')
    mh = ws[f'A{match_hdr_row}']
    mh.value = (f'▌ PROPERTY MATCHES  —  {len(matches)} properties matched across '
                f'{len(set(m[0] for m in matches))} clients  |  '
                f'Sorted by client then price')
    mh.font  = Font(name='Arial', bold=True, size=10, color=WHITE)
    mh.fill  = NAVY_FILL
    mh.alignment = LFT
    ws.row_dimensions[match_hdr_row].height = 20

    # ── Match column headers ──────────────────────────────────────────────────────
    match_col_hdr_row = match_hdr_row + 1
    ws.row_dimensions[match_col_hdr_row].height = 26
    for ci, (key, w, al) in enumerate(MATCH_COLS, 1):
        c = ws.cell(match_col_hdr_row, ci, key)
        c.font  = HDR_FONT
        c.fill  = TEAL2_FILL
        c.alignment = CTR
        c.border = BDR
        # Only set column width if wider than profile column width
        col_letter = get_column_letter(ci)
        existing_w = ws.column_dimensions[col_letter].width or 0
        if w > existing_w:
            ws.column_dimensions[col_letter].width = w

    # ── Match data rows ──────────────────────────────────────────────────────────
    if not matches:
        no_match_row = match_col_hdr_row + 1
        ws.merge_cells(f'A{no_match_row}:{get_column_letter(MC)}{no_match_row}')
        nm = ws[f'A{no_match_row}']
        nm.value = ('No matches found yet. Add client profiles above, then '
                    'save and re-run the daily update script.')
        nm.font  = HINT_FONT
        nm.fill  = TEAL_L_FILL
        nm.alignment = CTR
        ws.row_dimensions[no_match_row].height = 22
    else:
        # Sort: by client name order (matching client list order), then source, then price desc
        client_order = {c['name']: i for i, c in enumerate(clients)}
        def sort_key(m):
            cname, l = m
            ci = client_order.get(cname, 99)
            src_order = 0 if 'For Sale' in l['source'] else 1
            p = parse_price(l.get('price','')) or 0
            return (ci, src_order, -p)
        sorted_matches = sorted(matches, key=sort_key)

        prev_client = None
        for ri, (cname, listing) in enumerate(sorted_matches):
            rn = match_col_hdr_row + 1 + ri
            ws.row_dimensions[rn].height = 17

            # Alternating fill by client
            cidx = client_order.get(cname, 0)
            fill = BLU_FILL if cidx % 2 == 0 else GRN2_FILL

            src = listing['source']
            vals = [
                cname,
                src,
                listing.get('address',''),
                listing.get('suburb',''),
                listing.get('type',''),
                listing.get('beds',''),
                listing.get('baths',''),
                listing.get('price',''),
                listing.get('agency',''),
                listing.get('agents',''),
                listing.get('agent_mobile',''),
            ]

            for ci, v in enumerate(vals, 1):
                c = ws.cell(rn, ci, v)
                c.font   = ROW_FONT
                c.fill   = fill
                c.alignment = LFT if ci in (1,3,4,8,9,10) else CTR
                c.border = VBDR

            # Bold client name when it changes
            if cname != prev_client:
                ws.cell(rn, 1).font = BOLD_FONT
                prev_client = cname

        # Summary row at bottom
        sum_row = match_col_hdr_row + 1 + len(sorted_matches) + 1
        ws.merge_cells(f'A{sum_row}:{get_column_letter(MC)}{sum_row}')
        sr = ws[f'A{sum_row}']
        client_summary = ', '.join(
            f"{c['name']}: {sum(1 for m in matches if m[0]==c['name'])} matches"
            for c in clients if any(m[0]==c['name'] for m in matches)
        )
        sr.value = f'Summary: {client_summary}'
        sr.font  = Font(name='Arial', size=9, italic=True, color=TEAL)
        sr.fill  = TEAL_L_FILL
        sr.alignment = LFT
        ws.row_dimensions[sum_row].height = 18

    print(f"  '👥 Buyers' sheet built: {len(clients)} clients, {len(matches)} matches")

def run():
    print(f'Loading workbook {XL_PATH}...')
    wb = load_workbook(XL_PATH)
    print(f'  Existing sheets: {wb.sheetnames}')

    # 1. Read existing client profiles (preserves them)
    clients = read_client_profiles(wb)

    # 2. Save to JSON
    save_clients_json(clients)

    # 3. Load listings for matching
    forsale, offmarket = load_all_listings(BASE)

    # 4. Load agent phones
    phone_lookup = load_agent_phones(BASE)

    # 5. Enrich for-sale listings with agent mobile
    for l in forsale:
        names = [n.strip() for n in re.split(r'[,/]', l.get('agents',''))]
        mobiles = [phone_lookup.get(n.lower(),'') for n in names if phone_lookup.get(n.lower(),'')]
        l['agent_mobile'] = ' / '.join(m for m in mobiles[:2] if m)

    # 6. Find matches
    matches = find_matches(clients, forsale, offmarket)

    # 7. Build sheet
    build_sheet(wb, clients, matches)

    wb.save(XL_PATH)
    print(f'\nSaved → {XL_PATH}')

if __name__ == '__main__':
    run()

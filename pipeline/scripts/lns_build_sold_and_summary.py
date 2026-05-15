"""
build_sold_and_summary.py
  1. Parses sold listings from console-messages file
  2. Adds "🏡 Sold Properties" sheet to the Excel workbook
  3. Flags any current For-Sale listings that appear in sold → marks them SOLD
  4. Rebuilds the Summary sheet with a full 2026 LNS market snapshot
"""
import json, re, glob, sys
from pathlib import Path
from datetime import date, datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ── path discovery ─────────────────────────────────────────────────────────────
_cands = glob.glob('/sessions/*/mnt/Downloads/lns_agents_full.json')
BASE = Path(_cands[0]).parent if _cands else Path.home() / 'Downloads'
XL_PATH        = BASE / 'LNS_Real_Estate_Agents_FULL.xlsx'
AGENTS_JSON    = BASE / 'lns_agents_full.json'
FIRST_SEEN_FILE= BASE / 'lns_first_seen.json'
SOLD_SHEET     = "🏡 Sold Properties"
FORSALE_SHEET  = "🏠 Properties For Sale"
SUMMARY_SHEET  = "Summary"

# ── load agent phone map ───────────────────────────────────────────────────────
with open(AGENTS_JSON) as f:
    agents_raw = json.load(f)
agent_phone_map = {}
for a in agents_raw:
    key = (a.get('name') or '').strip().lower()
    if key:
        agent_phone_map[key] = {'mobile': a.get('mobile',''), 'telephone': a.get('telephone','')}

# ── load sold listings from JSON (written daily by scrape step) ───────────────
SOLD_FILE    = BASE / 'lns_sold_raw.json'
FORSALE_FILE = BASE / 'lns_listings_raw.json'

print(f"Reading sold listings from {SOLD_FILE}...")
with open(SOLD_FILE) as f:
    sold_listings = json.load(f)
print(f"  {len(sold_listings)} sold listings loaded")

# ── load for-sale listings ─────────────────────────────────────────────────────
with open(FORSALE_FILE) as f:
    forsale_listings = json.load(f)
forsale_ids = {str(l['id']) for l in forsale_listings}
sold_ids    = {str(l['id']) for l in sold_listings}
overlap     = forsale_ids & sold_ids
print(f"  {len(overlap)} for-sale listings now appear as sold (will be flagged)")

# ── helpers ────────────────────────────────────────────────────────────────────
THIN   = Side(style='thin',  color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR    = Alignment(horizontal='center', vertical='center')
LFT    = Alignment(horizontal='left',   vertical='center')
ROW_FONT  = Font(name='Arial', size=9)
BOLD_FONT = Font(name='Arial', size=9, bold=True)
HDR_FONT  = Font(name='Arial', size=10, bold=True, color='FFFFFF')
SUB_FONT  = Font(name='Arial', size=9,  bold=True, color='FFFFFF')
LINK_FONT = Font(name='Arial', size=9, color='1565C0', underline='single')
NAVY  = '1A3A5C'; TEAL = '0D7377'; WHITE = 'FFFFFF'
NAVY_FILL = PatternFill('solid', fgColor=NAVY)
TEAL_FILL = PatternFill('solid', fgColor=TEAL)
ALT_FILL  = PatternFill('solid', fgColor='EBF2FA')
SOLD_GOLD = PatternFill('solid', fgColor='FFF8E1')
GREEN_FILL= PatternFill('solid', fgColor='E8F5E9')
RED_FILL  = PatternFill('solid', fgColor='FFEBEE')

SUBURB_ORDER = [
    'Mosman','Neutral Bay','Cremorne','Cremorne Point','Kirribilli',
    'Mcmahons Point','Lavender Bay','Milsons Point','Cammeray',
    'Crows Nest','North Sydney','Waverton','Wollstonecraft','Kurraba Point',
    'Naremburn','St Leonards','Artarmon','Greenwich','Longueville'
]

def fmt_phone(p):
    if not p: return ''
    d = re.sub(r'\D', '', str(p))
    if len(d) == 10:
        return f"{d[:4]} {d[4:7]} {d[7:]}" if d.startswith('04') else f"{d[:2]} {d[2:6]} {d[6:]}"
    return str(p).strip()

def lookup_phones(names_str):
    if not names_str: return '', ''
    mob, pho = [], []
    for n in names_str.split('/'):
        info = agent_phone_map.get(n.strip().lower(), {})
        m = fmt_phone(info.get('mobile',''))
        t = fmt_phone(info.get('telephone',''))
        if m: mob.append(m)
        if t and t not in pho: pho.append(t)
    return ' / '.join(mob[:2]), ' / '.join(pho[:1])

# ── first-seen tracker ────────────────────────────────────────────────────────
def update_first_seen(listings):
    """Record the first date each listing ID was seen. Never overwrites existing dates."""
    today_iso = date.today().isoformat()
    first_seen = {}
    if FIRST_SEEN_FILE.exists():
        with open(FIRST_SEEN_FILE) as f:
            first_seen = json.load(f)
    new_count = 0
    for l in listings:
        lid = str(l.get('id',''))
        if lid and lid not in first_seen:
            first_seen[lid] = today_iso
            new_count += 1
    with open(FIRST_SEEN_FILE, 'w') as f:
        json.dump(first_seen, f)
    print(f"  First-seen tracker: {len(first_seen)} total, {new_count} new today")
    return first_seen

def parse_sold_date(date_str):
    """Parse a sold date string like '31 Mar 2026' → date object, or None."""
    if not date_str: return None
    for fmt in ('%d %b %Y', '%d %B %Y'):
        try: return datetime.strptime(date_str.strip(), fmt).date()
        except: pass
    return None

def parse_price(text):
    """Return best numeric price estimate from a price string, or 0."""
    if not text or 'withheld' in text.lower() or 'contact' in text.lower(): return 0
    nums = [float(n.replace(',','')) for n in re.findall(r'[\d,]+(?:\.\d+)?', text)
            if n.replace(',','') and float(n.replace(',','')) > 50000]
    if not nums: return 0
    return sum(nums) / len(nums)   # mid-point of range if two numbers

PRICE_BRACKETS = [
    (0,       1_000_000,  'Under $1M'),
    (1_000_000,1_500_000, '$1M – $1.5M'),
    (1_500_000,2_000_000, '$1.5M – $2M'),
    (2_000_000,3_000_000, '$2M – $3M'),
    (3_000_000,5_000_000, '$3M – $5M'),
    (5_000_000,999_999_999,'$5M+'),
]

def price_bracket(val):
    if val == 0: return 'Price Withheld'
    for lo, hi, label in PRICE_BRACKETS:
        if lo <= val < hi: return label
    return '$5M+'

def cell(ws, row, col, value, font=None, fill=None, align=None, border=True, hyperlink=None):
    c = ws.cell(row, col, value)
    c.font   = font   or ROW_FONT
    c.border = BORDER if border else None
    c.alignment = align or LFT
    if fill: c.fill = fill
    if hyperlink: c.hyperlink = hyperlink
    return c

# ═══════════════════════════════════════════════════════════════════════════════
# 1. BUILD SOLD PROPERTIES SHEET
# ═══════════════════════════════════════════════════════════════════════════════
SOLD_COLS = [
    ('Sold Date',    13, CTR),
    ('Method',       18, CTR),
    ('Address',      38, LFT),
    ('Suburb',       16, LFT),
    ('Type',         14, CTR),
    ('Beds',          6, CTR),
    ('Baths',         6, CTR),
    ('Cars',          6, CTR),
    ('Land (m²)',    10, CTR),
    ('Sold Price',   18, LFT),
    ('Agency',       28, LFT),
    ('Agents',       30, LFT),
    ('Agent Mobile', 16, CTR),
    ('Agent Phone',  14, CTR),
    ('Domain Link',  14, CTR),
]
N_SOLD = len(SOLD_COLS)

def build_sold_sheet(wb, today_str):
    if SOLD_SHEET in wb.sheetnames: del wb[SOLD_SHEET]
    # Insert after For Sale sheet
    idx = wb.sheetnames.index(FORSALE_SHEET) + 1 if FORSALE_SHEET in wb.sheetnames else 2
    ws = wb.create_sheet(SOLD_SHEET, idx)
    ws.freeze_panes = 'B3'

    # Title
    ws.merge_cells(f'A1:{get_column_letter(N_SOLD)}1')
    tc = ws['A1']
    tc.value = f'LNS Sold Properties (recent ~3 months)  |  domain.com.au  |  {len(sold_listings)} sales  |  Updated {today_str}'
    tc.font = Font(name='Arial', bold=True, size=11, color=WHITE)
    tc.fill = NAVY_FILL; tc.alignment = CTR
    ws.row_dimensions[1].height = 20

    # Headers
    ws.row_dimensions[2].height = 26
    for ci, (key, w, al) in enumerate(SOLD_COLS, 1):
        c = ws.cell(2, ci, key)
        c.font = HDR_FONT; c.fill = NAVY_FILL; c.alignment = CTR; c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Sort: suburb order then sold date descending
    suburb_rank = {s: i for i, s in enumerate(SUBURB_ORDER)}
    def sold_sort(l):
        dt = parse_sold_date(l.get('soldDate',''))
        ts = dt.toordinal() if dt else 0
        return (suburb_rank.get(l.get('primarySuburb',''), 99), -ts)

    sorted_sold = sorted(sold_listings, key=sold_sort)

    # Method colours
    method_fills = {
        'Auction':            PatternFill('solid', fgColor='E8EAF6'),
        'Private Treaty':     PatternFill('solid', fgColor='E8F5E9'),
        'Prior to Auction':   PatternFill('solid', fgColor='FFF8E1'),
        'Expression of Interest': PatternFill('solid', fgColor='E3F2FD'),
    }

    current_suburb = None; row_num = 3
    col_keys = [c[0] for c in SOLD_COLS]

    for l in sorted_sold:
        sub = l.get('primarySuburb', l.get('suburb',''))
        if sub != current_suburb:
            current_suburb = sub
            sub_count = sum(1 for x in sorted_sold if x.get('primarySuburb', x.get('suburb','')) == current_suburb)
            ws.merge_cells(f'A{row_num}:{get_column_letter(N_SOLD)}{row_num}')
            sc = ws.cell(row_num, 1)
            sc.value = f'  {current_suburb.upper()}  —  {sub_count} recent sales'
            sc.font = SUB_FONT; sc.fill = TEAL_FILL; sc.alignment = LFT
            ws.row_dimensions[row_num].height = 15; row_num += 1

        mob, pho = lookup_phones(l.get('agentNames',''))
        method = l.get('method','')
        row_fill = method_fills.get(method, ALT_FILL if row_num%2==0 else None)
        address = ', '.join(filter(None,[l.get('street',''),l.get('suburb',''),l.get('state',''),l.get('postcode','')]))
        url = l.get('url','')
        values = {
            'Sold Date':    l.get('soldDate',''),
            'Method':       method,
            'Address':      address,
            'Suburb':       sub,
            'Type':         l.get('propertyType',''),
            'Beds':         l.get('beds',0) or '',
            'Baths':        l.get('baths',0) or '',
            'Cars':         l.get('parking',0) or '',
            'Land (m²)':   l.get('landSize',0) or '',
            'Sold Price':   l.get('soldPrice',''),
            'Agency':       l.get('agencyName',''),
            'Agents':       l.get('agentNames',''),
            'Agent Mobile': mob,
            'Agent Phone':  pho,
            'Domain Link':  url,
        }
        for ci, key in enumerate(col_keys, 1):
            val = values.get(key,'')
            c = ws.cell(row_num, ci, val)
            c.border = BORDER; c.font = ROW_FONT; c.alignment = SOLD_COLS[ci-1][2]
            if row_fill: c.fill = row_fill
            if key == 'Domain Link' and val:
                short = '…'+val.split('domain.com.au')[-1][-28:]
                c.value = short; c.hyperlink = val; c.font = LINK_FONT
            elif key == 'Sold Price':
                c.font = Font(name='Arial', bold=True, size=9,
                    color='2E7D32' if '$' in str(val) else '888888')
            elif key == 'Method':
                mf = {'Auction':'283593','Private Treaty':'2E7D32',
                      'Prior to Auction':'E65100','Expression of Interest':'1565C0'}
                c.font = Font(name='Arial', bold=True, size=9, color=mf.get(method,'000000'))
            elif key in ('Agent Mobile','Agent Phone') and val:
                c.font = Font(name='Arial', size=9, color='1565C0')
        ws.row_dimensions[row_num].height = 15; row_num += 1

    # Footer
    row_num += 1
    ws.merge_cells(f'A{row_num}:{get_column_letter(N_SOLD)}{row_num}')
    ws.cell(row_num,1).value = (
        f'Total: {len(sold_listings)} recent sold  |  '
        f'Private Treaty: {sum(1 for l in sold_listings if l.get("method")=="Private Treaty")}  |  '
        f'Auction: {sum(1 for l in sold_listings if l.get("method")=="Auction")}  |  '
        f'Prior to Auction: {sum(1 for l in sold_listings if l.get("method")=="Prior to Auction")}  |  '
        f'Price Withheld: {sum(1 for l in sold_listings if "withheld" in (l.get("soldPrice","")).lower())}  |  '
        f'Source: domain.com.au'
    )
    ws.cell(row_num,1).font = Font(name='Arial', italic=True, size=9, color='555555')
    ws.cell(row_num,1).alignment = CTR
    print(f"  '{SOLD_SHEET}' sheet built: {len(sold_listings)} listings")

# ═══════════════════════════════════════════════════════════════════════════════
# 2. FLAG SOLD PROPERTIES IN FOR-SALE SHEET
# ═══════════════════════════════════════════════════════════════════════════════
def flag_sold_in_forsale_sheet(wb):
    if FORSALE_SHEET not in wb.sheetnames:
        print("  For Sale sheet not found, skipping flag step")
        return
    ws = wb[FORSALE_SHEET]
    sold_flag_fill = PatternFill('solid', fgColor='FFCDD2')
    flagged = 0

    # Build sold address set for fallback matching (normalised lowercase, digits+letters only)
    def norm_addr(s): return re.sub(r'[^a-z0-9]', '', str(s).lower())
    sold_addresses = {norm_addr(l.get('address','')) for l in sold_listings if l.get('address')}

    headers = {ws.cell(2, c).value: c for c in range(1, ws.max_column+1)}
    status_col  = headers.get('Status', 1)
    domain_col  = headers.get('Domain Link', None)
    addr_col    = headers.get('Address', None)

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        matched = False
        # Primary: match by listing ID extracted from hyperlink URL
        if domain_col:
            link_cell = row[domain_col-1]
            url = str(link_cell.hyperlink.target if link_cell.hyperlink else link_cell.value or '')
            m = re.search(r'-(\d{7,12})(?:/|$)', url)
            if m and m.group(1) in sold_ids:
                matched = True
        # Fallback: match by address string
        if not matched and addr_col:
            addr_val = str(row[addr_col-1].value or '')
            if addr_val and norm_addr(addr_val) in sold_addresses:
                matched = True
        if matched:
            for c in row:
                c.fill = sold_flag_fill
            status_cell = row[status_col-1]
            status_cell.value = '✅ SOLD'
            status_cell.font = Font(name='Arial', bold=True, size=9, color='B71C1C')
            flagged += 1
    print(f"  Flagged {flagged} listings in For Sale sheet as SOLD")

# ═══════════════════════════════════════════════════════════════════════════════
# 3. REBUILD SUMMARY SHEET WITH 2026 MARKET SNAPSHOT
# ═══════════════════════════════════════════════════════════════════════════════
def rebuild_summary(wb, today_str, first_seen):
    if SUMMARY_SHEET in wb.sheetnames: del wb[SUMMARY_SHEET]
    # Re-insert at last position
    ws = wb.create_sheet(SUMMARY_SHEET)

    # Column widths
    COL_W = {'A':22,'B':14,'C':14,'D':14,'E':14,'F':14,'G':14,'H':14,'I':14,'J':14,'K':14,'L':14,'M':14}
    for col, w in COL_W.items():
        ws.column_dimensions[col].width = w

    today = date.today()
    week_ago  = today - timedelta(days=7)
    month_str = today.strftime('%B %Y')   # e.g. "April 2026"
    month_start = today.replace(day=1)

    # ── price data for all listings ──────────────────────────────────────────
    # For-sale: parse price
    forsale_by_suburb  = {s:[] for s in SUBURB_ORDER}
    forsale_priced     = {s:[] for s in SUBURB_ORDER}
    for l in forsale_listings:
        s = l.get('primarySuburb', l.get('suburb',''))
        if s in forsale_by_suburb:
            forsale_by_suburb[s].append(l)
            p = parse_price(l.get('price',''))
            if p: forsale_priced[s].append(p)

    # Sold: parse price
    sold_by_suburb  = {s:[] for s in SUBURB_ORDER}
    sold_priced     = {s:[] for s in SUBURB_ORDER}
    for l in sold_listings:
        s = l.get('primarySuburb', l.get('suburb',''))
        if s in sold_by_suburb:
            sold_by_suburb[s].append(l)
            p = parse_price(l.get('soldPrice',''))
            if p: sold_priced[s].append(p)

    # Sale methods
    def method_counts(listings):
        mc = {}
        for l in listings:
            m = l.get('method','Other')
            mc[m] = mc.get(m,0)+1
        return mc

    BRACKETS = [b[2] for b in PRICE_BRACKETS] + ['Price Withheld']
    def bracket_counts(prices_list, all_list):
        bc = {b:0 for b in BRACKETS}
        priced_ids = {id(l) for l in all_list if parse_price(l.get('soldPrice','') or l.get('price',''))}
        for l in all_list:
            p = parse_price(l.get('soldPrice','') or l.get('price',''))
            if p == 0:
                bc['Price Withheld'] += 1
            else:
                bc[price_bracket(p)] += 1
        return bc

    row = 1

    # ── TITLE ──────────────────────────────────────────────────────────────────
    def title_row(text, fill, sz=12, span_cols=13):
        nonlocal row
        ws.merge_cells(f'A{row}:{get_column_letter(span_cols)}{row}')
        c = ws.cell(row,1,text)
        c.font = Font(name='Arial',bold=True,size=sz,color='FFFFFF')
        c.fill = fill; c.alignment = CTR
        ws.row_dimensions[row].height = 22; row += 1

    def section_title(text, fill=None):
        nonlocal row
        fill = fill or TEAL_FILL
        ws.merge_cells(f'A{row}:M{row}')
        c = ws.cell(row,1,text)
        c.font = Font(name='Arial',bold=True,size=10,color='FFFFFF')
        c.fill = fill; c.alignment = LFT
        ws.row_dimensions[row].height = 18; row += 1

    def hdr(r,c_num,text,fill=None,align=CTR):
        c = ws.cell(r,c_num,text)
        c.font = HDR_FONT; c.fill = fill or NAVY_FILL; c.alignment = align; c.border=BORDER
        return c

    def val(r,c_num,v,bold=False,fill=None,align=CTR,color='000000',bdr=True):
        c = ws.cell(r,c_num,v)
        c.font = Font(name='Arial',bold=bold,size=9,color=color)
        c.alignment = align
        if fill: c.fill = fill
        if bdr: c.border = BORDER
        return c

    # ─────────────────────────────────────────────────────────────────────────
    title_row(f'Lower North Shore Real Estate  —  2026 Market Snapshot  |  Updated {today_str}',
              NAVY_FILL, sz=13)
    total_fs  = sum(len(v) for v in forsale_by_suburb.values())
    total_sol = sum(len(v) for v in sold_by_suburb.values())
    row += 1  # spacer

    # ── NEW LISTINGS / SALES — THIS WEEK & THIS MONTH ─────────────────────────
    week_new_by_suburb  = {s:0 for s in SUBURB_ORDER}
    month_new_by_suburb = {s:0 for s in SUBURB_ORDER}
    for l in forsale_listings:
        s = l.get('primarySuburb', l.get('suburb',''))
        if s not in week_new_by_suburb: continue
        lid = str(l.get('id',''))
        fs_date_str = first_seen.get(lid)
        if not fs_date_str: continue
        try: fs_date = date.fromisoformat(fs_date_str)
        except: continue
        if fs_date >= week_ago:    week_new_by_suburb[s]  += 1
        if fs_date >= month_start: month_new_by_suburb[s] += 1

    week_sold_by_suburb  = {s:0 for s in SUBURB_ORDER}
    month_sold_by_suburb = {s:0 for s in SUBURB_ORDER}
    for l in sold_listings:
        s = l.get('primarySuburb', l.get('suburb',''))
        if s not in week_sold_by_suburb: continue
        sd = parse_sold_date(l.get('soldDate',''))
        if not sd: continue
        if sd >= week_ago:    week_sold_by_suburb[s]  += 1
        if sd >= month_start: month_sold_by_suburb[s] += 1

    AMBER_HDR   = PatternFill('solid', fgColor='E65100')
    GREEN_HDR   = PatternFill('solid', fgColor='2E7D32')
    AMBER_LIGHT = PatternFill('solid', fgColor='FFF3E0')
    GREEN_LIGHT = PatternFill('solid', fgColor='E8F5E9')

    def snapshot_table(label, new_by_sub, sold_by_sub, hdr_fill, alt_fill):
        nonlocal row
        section_title(f'  {label}', hdr_fill)
        hdr(row,1,'Suburb',       fill=hdr_fill)
        hdr(row,2,'New Listings', fill=hdr_fill)
        hdr(row,3,'Sold',         fill=hdr_fill)
        hdr(row,4,'Note',         fill=hdr_fill)
        ws.row_dimensions[row].height = 22; row += 1
        total_new = total_sold_snap = 0
        for i, s in enumerate(SUBURB_ORDER):
            n  = new_by_sub[s]
            so = sold_by_sub[s]
            total_new       += n
            total_sold_snap += so
            if n == 0 and so == 0: continue
            alt = alt_fill if i%2==0 else None
            if n > 0 and so > 0:   note = f'{n} new listing{"s" if n>1 else ""}  ·  {so} sold'
            elif n > 0:            note = f'{n} new listing{"s" if n>1 else ""}'
            else:                  note = f'{so} sold'
            val(row,1,s,   bold=True,fill=alt,align=LFT)
            val(row,2,n,   bold=True,fill=alt,color='1A3A5C' if n  else '888888')
            val(row,3,so,  bold=True,fill=alt,color='2E7D32' if so else '888888')
            val(row,4,note,fill=alt,align=LFT,color='555555')
            ws.row_dimensions[row].height = 15; row += 1
        val(row,1,'TOTAL LNS',bold=True,fill=hdr_fill,align=LFT,color='FFFFFF')
        val(row,2,total_new,       bold=True,fill=hdr_fill,color='FFFFFF')
        val(row,3,total_sold_snap, bold=True,fill=hdr_fill,color='FFFFFF')
        val(row,4,'',fill=hdr_fill)
        ws.row_dimensions[row].height = 16; row += 2

    snapshot_table(
        f'THIS WEEK  —  New to Market & Sold  (last 7 days, as at {today_str})',
        week_new_by_suburb, week_sold_by_suburb,
        AMBER_HDR, AMBER_LIGHT
    )
    snapshot_table(
        f'THIS MONTH  —  {month_str}  |  New Listings & Sales',
        month_new_by_suburb, month_sold_by_suburb,
        GREEN_HDR, GREEN_LIGHT
    )

    # ── SECTION 1: Suburb Overview ────────────────────────────────────────────
    section_title('  SUBURB OVERVIEW  —  For Sale vs Recent Sold (last ~3 months)')
    hdr(row,1,'Suburb')
    hdr(row,2,'For Sale')
    hdr(row,3,'Recently Sold')
    hdr(row,4,'Sale Rate')
    hdr(row,5,'Avg Ask Price')
    hdr(row,6,'Avg Sold Price')
    hdr(row,7,'Auction %')
    hdr(row,8,'Private Treaty %')
    hdr(row,9,'Price Withheld %')
    ws.column_dimensions['A'].width = 22
    ws.row_dimensions[row].height = 26; row += 1

    for i, s in enumerate(SUBURB_ORDER):
        fs   = forsale_by_suburb[s]
        sol  = sold_by_suburb[s]
        fsp  = forsale_priced[s]
        solp = sold_priced[s]
        alt  = ALT_FILL if i%2==0 else None

        sale_rate = f'{len(sol)/len(fs)*100:.0f}%' if fs else '-'
        avg_ask   = f'${sum(fsp)/len(fsp)/1e6:.2f}M' if fsp else '-'
        avg_sold  = f'${sum(solp)/len(solp)/1e6:.2f}M' if solp else '-'

        mc = method_counts(sol)
        total_sol_s = len(sol) or 1
        auc_pct  = f'{mc.get("Auction",0)/total_sol_s*100:.0f}%'
        pt_pct   = f'{mc.get("Private Treaty",0)/total_sol_s*100:.0f}%'
        pw_pct   = f'{sum(1 for l in sol if "withheld" in l.get("soldPrice","").lower())/total_sol_s*100:.0f}%'

        val(row,1,s,    bold=True,fill=alt,align=LFT)
        val(row,2,len(fs),  fill=alt,color='1A3A5C',bold=True)
        val(row,3,len(sol), fill=alt,color='0D7377',bold=True)
        val(row,4,sale_rate,fill=alt,color='2E7D32' if len(sol)>len(fs)*0.3 else '000000')
        val(row,5,avg_ask,  fill=alt,align=LFT)
        val(row,6,avg_sold, fill=alt,align=LFT,color='2E7D32' if solp else '888888')
        val(row,7,auc_pct,  fill=alt)
        val(row,8,pt_pct,   fill=alt)
        val(row,9,pw_pct,   fill=alt,color='888888')
        ws.row_dimensions[row].height = 15; row += 1

    # Totals row
    total_fsp  = [p for v in forsale_priced.values() for p in v]
    total_solp = [p for v in sold_priced.values() for p in v]
    val(row,1,'TOTAL LNS',bold=True,fill=NAVY_FILL,align=LFT,color='FFFFFF')
    val(row,2,total_fs, bold=True,fill=NAVY_FILL,color='FFFFFF')
    val(row,3,total_sol,bold=True,fill=NAVY_FILL,color='FFFFFF')
    val(row,4,f'{total_sol/total_fs*100:.0f}%' if total_fs else '-',fill=NAVY_FILL,color='FFFFFF')
    val(row,5,f'${sum(total_fsp)/len(total_fsp)/1e6:.2f}M' if total_fsp else '-',fill=NAVY_FILL,color='FFFFFF',align=LFT)
    val(row,6,f'${sum(total_solp)/len(total_solp)/1e6:.2f}M' if total_solp else '-',fill=NAVY_FILL,color='FFFFFF',align=LFT)
    ws.row_dimensions[row].height = 16; row += 2

    # ── SECTION 2: Price Bracket Matrix — For Sale ────────────────────────────
    BRACKET_LABELS = [b[2] for b in PRICE_BRACKETS]
    N_BRACKETS = len(BRACKET_LABELS)

    def price_matrix(title, listings_by_suburb, price_key, fill_hdr):
        nonlocal row
        section_title(f'  {title}', fill_hdr)

        # Header row
        hdr(row,1,'Suburb',fill=fill_hdr)
        hdr(row,2,'Total',fill=fill_hdr)
        for j,label in enumerate(BRACKET_LABELS):
            hdr(row, 3+j, label, fill=fill_hdr)
        hdr(row, 3+N_BRACKETS, 'Price\nWithheld', fill=fill_hdr)
        ws.row_dimensions[row].height = 28; row += 1

        grand = {b:0 for b in BRACKET_LABELS+['Price Withheld']}
        for i,s in enumerate(SUBURB_ORDER):
            listings = listings_by_suburb[s]
            if not listings: continue
            alt = ALT_FILL if i%2==0 else None
            bc  = {b:0 for b in BRACKET_LABELS+['Price Withheld']}
            for l in listings:
                p = parse_price(l.get(price_key,''))
                b = price_bracket(p) if p else 'Price Withheld'
                if b in bc: bc[b] += 1
                if b in grand: grand[b] += 1

            val(row,1,s,bold=True,fill=alt,align=LFT)
            val(row,2,len(listings),bold=True,fill=alt)
            for j,label in enumerate(BRACKET_LABELS):
                n = bc[label]
                val(row,3+j,n if n else '',fill=alt,
                    color='1A3A5C' if n>=5 else ('0D7377' if n>=2 else '888888'))
            val(row,3+N_BRACKETS,bc['Price Withheld'] or '',fill=alt,color='AAAAAA')
            ws.row_dimensions[row].height = 15; row += 1

        # Grand total
        val(row,1,'TOTAL LNS',bold=True,fill=fill_hdr,align=LFT,color='FFFFFF')
        val(row,2,sum(grand.values()),bold=True,fill=fill_hdr,color='FFFFFF')
        for j,label in enumerate(BRACKET_LABELS):
            val(row,3+j,grand[label],bold=True,fill=fill_hdr,color='FFFFFF')
        val(row,3+N_BRACKETS,grand['Price Withheld'],bold=True,fill=fill_hdr,color='FFFFFF')
        ws.row_dimensions[row].height = 16; row += 2

    price_matrix(
        'PRICE BRACKET BREAKDOWN — Properties Currently For Sale',
        forsale_by_suburb, 'price',
        TEAL_FILL
    )
    price_matrix(
        'PRICE BRACKET BREAKDOWN — Recently Sold (Price Disclosed)',
        sold_by_suburb, 'soldPrice',
        PatternFill('solid', fgColor='2E7D32')
    )

    # ── SECTION 3: Agent leaderboard (most sold recently) ────────────────────
    section_title('  TOP SELLING AGENTS  —  Most Recent Sales (last ~3 months)', NAVY_FILL)
    hdr(row,1,'Agent')
    hdr(row,2,'Agency',fill=NAVY_FILL)
    hdr(row,3,'Sales (recent)',fill=NAVY_FILL)
    hdr(row,4,'Auction',fill=NAVY_FILL)
    hdr(row,5,'Private Treaty',fill=NAVY_FILL)
    hdr(row,6,'Primary Suburb',fill=NAVY_FILL)
    ws.row_dimensions[row].height = 26; row += 1

    agent_sales = {}
    for l in sold_listings:
        for name in (l.get('agentNames','') or '').split('/'):
            name = name.strip()
            if not name: continue
            if name not in agent_sales:
                agent_sales[name] = {'count':0,'auction':0,'pt':0,'agency':l.get('agencyName',''),'suburb':l.get('primarySuburb','')}
            agent_sales[name]['count'] += 1
            if l.get('method') == 'Auction': agent_sales[name]['auction'] += 1
            elif l.get('method') == 'Private Treaty': agent_sales[name]['pt'] += 1

    top_agents = sorted(agent_sales.items(), key=lambda x:-x[1]['count'])[:30]
    for i,(name,data) in enumerate(top_agents):
        alt = ALT_FILL if i%2==0 else None
        val(row,1,name,bold=True,fill=alt,align=LFT)
        val(row,2,data['agency'],fill=alt,align=LFT)
        val(row,3,data['count'],bold=True,fill=alt,color='0D7377')
        val(row,4,data['auction'],fill=alt)
        val(row,5,data['pt'],fill=alt)
        val(row,6,data['suburb'],fill=alt,align=LFT)
        ws.row_dimensions[row].height = 15; row += 1

    # ── SECTION 4: Method breakdown per suburb ────────────────────────────────
    row += 1
    section_title('  SALE METHOD BREAKDOWN  —  Recent Sold', TEAL_FILL)
    hdr(row,1,'Suburb',fill=TEAL_FILL)
    hdr(row,2,'Total Sold',fill=TEAL_FILL)
    hdr(row,3,'Auction',fill=TEAL_FILL)
    hdr(row,4,'Auction %',fill=TEAL_FILL)
    hdr(row,5,'Private Treaty',fill=TEAL_FILL)
    hdr(row,6,'PT %',fill=TEAL_FILL)
    hdr(row,7,'Prior to Auction',fill=TEAL_FILL)
    hdr(row,8,'Price Withheld',fill=TEAL_FILL)
    ws.row_dimensions[row].height = 26; row += 1

    for i,s in enumerate(SUBURB_ORDER):
        sol = sold_by_suburb[s]
        if not sol: continue
        alt = ALT_FILL if i%2==0 else None
        mc  = method_counts(sol)
        total = len(sol) or 1
        pw  = sum(1 for l in sol if 'withheld' in l.get('soldPrice','').lower())
        auc = mc.get('Auction',0)
        pt  = mc.get('Private Treaty',0)
        pta = mc.get('Prior to Auction',0)
        val(row,1,s,bold=True,fill=alt,align=LFT)
        val(row,2,len(sol),bold=True,fill=alt)
        val(row,3,auc,fill=alt,color='283593' if auc else '888888')
        val(row,4,f'{auc/total*100:.0f}%',fill=alt)
        val(row,5,pt, fill=alt,color='2E7D32' if pt  else '888888')
        val(row,6,f'{pt/total*100:.0f}%',fill=alt)
        val(row,7,pta,fill=alt,color='E65100' if pta else '888888')
        val(row,8,pw, fill=alt,color='AAAAAA')
        ws.row_dimensions[row].height = 15; row += 1

    print(f"  '{SUMMARY_SHEET}' rebuilt with market snapshot")

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
today_str = date.today().strftime('%-d %b %Y')
print(f"\nLoading workbook {XL_PATH}...")
wb = load_workbook(XL_PATH)
print(f"  Existing sheets: {wb.sheetnames}")

first_seen = update_first_seen(forsale_listings)
build_sold_sheet(wb, today_str)
flag_sold_in_forsale_sheet(wb)
rebuild_summary(wb, today_str, first_seen)

wb.save(XL_PATH)
print(f"\nSaved → {XL_PATH}")
print(f"  Sheets: {wb.sheetnames}")
